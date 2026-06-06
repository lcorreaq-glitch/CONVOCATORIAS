"""KRINOS - Rankings + Desempates + Actas PDF + Reportes + Dashboard + Auditoría."""
import uuid
import io
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors as rl_colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from db import get_db, now_iso
from auth import get_current_user, require_roles, audit
from fastapi import Body

router = APIRouter(prefix="/api", tags=["reports"])


# Cupos por defecto INC2026 (Iniciativas Antioquia 2026)
DEFAULT_CUPOS_INC2026_SUBREGION = {
    "Urabá": 14, "Oriente": 10, "Occidente": 10, "Magdalena Medio": 6,
    "Bajo Cauca": 10, "Suroeste": 9, "Nordeste": 8, "Valle de Aburrá": 4, "Norte": 9,
}


# ==================== RANKING ====================
async def _compute_propuesta_score(db, convocatoria_id: str, propuesta_id: str, modo: str = "colectivo"):
    """Returns dict with puntaje_total and criterios_detalle.

    Modo ESTRICTO: la fuente del puntaje debe coincidir exactamente con el modo seleccionado.
    - 'colectivo'  -> SOLO evaluaciones_colectivas en estado Cerrada/Firmada (NO Abierta ni Reabierta).
                      Si hay varias (caso raro de reasignación de terna), toma la más reciente cerrada.
    - 'individual' -> SOLO promedio de evaluaciones_individuales en estado Finalizada/Firmada
                      (excluye Borrador y Reabierta para no contar puntajes en revisión).
    """
    if modo == "colectivo":
        # Solo evaluaciones efectivamente cerradas: no contamos las abiertas/reabiertas (puntajes en revisión)
        evs = await db.evaluaciones_colectivas.find({
            "propuesta_id": propuesta_id,
            "estado": {"$in": ["Cerrada", "Firmada"]},
        }).to_list(20)
        if not evs:
            return {"puntaje_total": 0, "puntaje_diferencial": 0, "criterios": {}, "fuente": "ninguna"}
        # Si hay varias (caso raro: terna reasignada), priorizar Firmada y la más reciente.
        evs.sort(key=lambda e: (e.get("firmada_at") or e.get("cerrada_at") or e.get("updated_at") or ""), reverse=True)
        firmadas = [e for e in evs if e.get("estado") == "Firmada"]
        ev = firmadas[0] if firmadas else evs[0]
        return {
            "puntaje_total": ev.get("puntaje_final", 0),
            "puntaje_diferencial": ev.get("puntaje_diferencial_total", 0),
            "criterios": ev.get("puntajes", {}),
            "fuente": "colectiva",
        }

    # modo == "individual" (o cualquier otro distinto a colectivo) -> usa SOLO individuales activas
    individuales = await db.evaluaciones_individuales.find({
        "propuesta_id": propuesta_id,
        "estado": {"$in": ["Finalizada", "Firmada"]},
        "etapa": {"$ne": "colectiva"},  # Excluir V2 de la etapa colectiva (no son individuales reales)
    }).to_list(50)
    if not individuales:
        return {"puntaje_total": 0, "puntaje_diferencial": 0, "criterios": {}, "fuente": "ninguna"}
    crit_sum, crit_cnt = {}, {}
    for ev in individuales:
        for cid, v in (ev.get("puntajes") or {}).items():
            try: vf = float(v)
            except Exception: continue
            crit_sum[cid] = crit_sum.get(cid, 0) + vf
            crit_cnt[cid] = crit_cnt.get(cid, 0) + 1
    promedio = {cid: round(crit_sum[cid] / crit_cnt[cid], 2) for cid in crit_sum}
    criterios = await db.criterios.find({"convocatoria_id": convocatoria_id}).to_list(100)
    total_of = sum(promedio.get(c["id"], 0) for c in criterios if c.get("oficial", True))
    total_dif = sum(promedio.get(c["id"], 0) for c in criterios if not c.get("oficial", True))
    return {
        "puntaje_total": round(total_of, 2),
        "puntaje_diferencial": round(total_dif, 2),
        "criterios": promedio,
        "fuente": "promedio_individuales",
    }


def _resolve_desempate(a: dict, b: dict, desempates: list, criterios_by_nombre: dict):
    """Returns negative if a should be ranked before b, positive otherwise, and the regla applied."""
    for d in desempates:
        campo = d["campo"]
        tipo = d["tipo_comparacion"]
        va, vb = None, None
        if campo.startswith("criterio:"):
            nombre = campo.split(":", 1)[1].strip()
            c = criterios_by_nombre.get(nombre)
            if c:
                va = a["criterios_detalle"].get(c["id"], 0)
                vb = b["criterios_detalle"].get(c["id"], 0)
        elif campo == "fecha_radicacion":
            # Combinar fecha + hora para desempate cronológico exacto
            # (las propuestas tienen `datos.fecha_radicacion` y `datos.hora_radicacion` por separado)
            fa = (a["datos"].get("fecha_radicacion") or "9999-12-31").split("T")[0]
            fb = (b["datos"].get("fecha_radicacion") or "9999-12-31").split("T")[0]
            ha = a["datos"].get("hora_radicacion") or "23:59:59"
            hb = b["datos"].get("hora_radicacion") or "23:59:59"
            va = f"{fa} {ha}"
            vb = f"{fb} {hb}"
        elif campo == "hora_radicacion":
            va = a["datos"].get("hora_radicacion") or "99:99:99"
            vb = b["datos"].get("hora_radicacion") or "99:99:99"
        elif campo == "sorteo":
            return (0, d["nombre"])
        if va is None or vb is None:
            continue
        if tipo == "mayor_valor":
            if va != vb:
                return ((vb - va), d["nombre"]) if isinstance(va, (int, float)) else (0, d["nombre"])
        elif tipo == "menor_valor":
            if va != vb:
                return ((va - vb), d["nombre"])
        elif tipo in ("fecha_mas_antigua", "hora_mas_antigua"):
            if va != vb:
                return ((-1 if va < vb else 1), d["nombre"])
        elif tipo in ("fecha_mas_reciente", "hora_mas_reciente"):
            if va != vb:
                return ((-1 if va > vb else 1), d["nombre"])
    return (0, None)


@router.post("/rankings/generar")
async def generar_ranking(convocatoria_id: str, agrupar_por: str = "subregion",
                          modo: str = "colectivo",
                          user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    propuestas = await db.propuestas.find({
        "convocatoria_id": convocatoria_id,
        "estado": {"$nin": ["Anulada", "No habilitada"]}
    }).to_list(5000)
    if not propuestas:
        raise HTTPException(status_code=400, detail="No hay propuestas habilitadas para rankear")

    # Get scores
    enriched = []
    for p in propuestas:
        score = await _compute_propuesta_score(db, convocatoria_id, p["id"], modo=modo)
        enriched.append({
            "propuesta_id": p["id"],
            "codigo": p.get("codigo"),
            "nombre": p.get("nombre"),
            "organizacion": p.get("organizacion"),
            "datos": p.get("datos", {}),
            "puntaje_total": score["puntaje_total"],
            "puntaje_diferencial": score["puntaje_diferencial"],
            "criterios_detalle": score["criterios"],
            "fuente": score["fuente"],
        })

    # Group
    grupos = {}
    for e in enriched:
        g = e["datos"].get(agrupar_por) or "Sin grupo"
        grupos.setdefault(g, []).append(e)

    desempates = await db.desempates.find({"convocatoria_id": convocatoria_id, "activo": True}).sort("orden", 1).to_list(50)
    criterios = await db.criterios.find({"convocatoria_id": convocatoria_id}).to_list(100)
    criterios_by_nombre = {c["nombre"]: c for c in criterios}

    # Sort each group
    import functools
    def cmp(a, b):
        if a["puntaje_total"] != b["puntaje_total"]:
            return -1 if a["puntaje_total"] > b["puntaje_total"] else 1
        # Tie-break
        result, regla = _resolve_desempate(a, b, desempates, criterios_by_nombre)
        if regla:
            a.setdefault("desempate_regla", regla)
            b.setdefault("desempate_regla", regla)
        return -1 if result < 0 else (1 if result > 0 else 0)

    resultado = {"grupos": [], "agrupacion": agrupar_por, "modo": modo,
                 "fecha_generacion": now_iso(), "convocatoria_id": convocatoria_id,
                 "id": str(uuid.uuid4()), "estado": "Preliminar"}

    # Cupos de ganadores por grupo (subregión, línea, etc.)
    # Se configura en convocatoria.configuracion.cupos_ganadores[<agrupacion>] = { grupo_nombre: cantidad }
    # Ejemplo INC2026: configuracion.cupos_ganadores.subregion = {"Urabá":14, "Oriente":10, ...}
    conv_doc = await db.convocatorias.find_one({"id": convocatoria_id}, {"_id": 0, "configuracion": 1})
    cupos_cfg = ((conv_doc or {}).get("configuracion") or {}).get("cupos_ganadores") or {}
    cupos_grupo = cupos_cfg.get(agrupar_por) or {}
    incentivos_no_asignados = []  # informe de cupos sobrantes
    total_cupos_configurados = 0
    total_ganadores_asignados = 0

    for g, items in sorted(grupos.items()):
        items.sort(key=functools.cmp_to_key(cmp))
        cupo_grupo = int(cupos_grupo.get(g, 0))
        if cupo_grupo:
            total_cupos_configurados += cupo_grupo
        for pos, it in enumerate(items, start=1):
            it["puesto"] = pos
            if cupo_grupo:
                if pos <= cupo_grupo:
                    it["resultado"] = "ganador"
                else:
                    it["resultado"] = "lista_espera"
            else:
                it["resultado"] = "ganador" if pos == 1 else "elegible"
        # Cupos vs propuestas reales del grupo
        ganadores_en_grupo = sum(1 for it in items if it.get("resultado") == "ganador")
        total_ganadores_asignados += ganadores_en_grupo
        sobrantes = cupo_grupo - ganadores_en_grupo if cupo_grupo else 0
        if sobrantes > 0:
            incentivos_no_asignados.append({
                "grupo": g, "cupo_configurado": cupo_grupo,
                "propuestas_disponibles": len(items),
                "ganadores_asignados": ganadores_en_grupo,
                "incentivos_sobrantes": sobrantes,
            })
        resultado["grupos"].append({
            "grupo": g, "items": items, "total": len(items),
            "cupo_ganadores": cupo_grupo if cupo_grupo else None,
            "ganadores_asignados": ganadores_en_grupo if cupo_grupo else None,
        })

    resultado["incentivos_no_asignados"] = incentivos_no_asignados
    resultado["total_cupos_configurados"] = total_cupos_configurados or None
    resultado["total_ganadores_asignados"] = total_ganadores_asignados if total_cupos_configurados else None
    resultado["total_incentivos_sobrantes"] = sum(x["incentivos_sobrantes"] for x in incentivos_no_asignados) or 0

    # Resumen de cobertura: cuántas propuestas no tienen fuente de puntaje válida en el modo seleccionado.
    sin_fuente = [e for e in enriched if e.get("fuente") == "ninguna"]
    resultado["cobertura"] = {
        "total_propuestas": len(enriched),
        "con_puntaje": len(enriched) - len(sin_fuente),
        "sin_puntaje": len(sin_fuente),
        "propuestas_sin_puntaje": [{"codigo": p.get("codigo"), "nombre": p.get("nombre"), "propuesta_id": p.get("propuesta_id")} for p in sin_fuente[:50]],
    }

    await db.rankings.insert_one(resultado)
    resultado.pop("_id", None)
    await audit(user, "generate", "rankings", resultado["id"], detalle=f"agrupar_por={agrupar_por}, modo={modo}")
    return resultado


@router.get("/rankings")
async def list_rankings(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    items = await db.rankings.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("fecha_generacion", -1).to_list(50)
    return items


@router.get("/rankings/{rid}")
async def get_ranking(rid: str, user: dict = Depends(get_current_user)):
    db = get_db()
    item = await db.rankings.find_one({"id": rid}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Ranking no encontrado")
    return item


@router.get("/rankings/{rid}/excel")
async def export_ranking_excel(rid: str, user: dict = Depends(get_current_user)):
    """Descarga el ranking como Excel con una hoja por grupo (línea/subregión/etc)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    db = get_db()
    rk = await db.rankings.find_one({"id": rid}, {"_id": 0})
    if not rk:
        raise HTTPException(status_code=404, detail="Ranking no encontrado")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="14776A")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["#", "Código", "Nombre propuesta", "Organización", "Municipio",
               "Subregión", "Puntaje total", "Bono priorización", "Puntaje final",
               "Puntaje diferencial", "Fuente", "Estado", "Cupos", "Desempate aplicado"]
    # Hoja Resumen
    wsr = wb.create_sheet("Resumen")
    wsr.append(["Convocatoria", rk.get("convocatoria_id", "")])
    wsr.append(["Modo", rk.get("modo", "")])
    wsr.append(["Agrupar por", rk.get("agrupar_por", "")])
    wsr.append(["Generado", rk.get("fecha_generacion", "")])
    cob = rk.get("cobertura") or {}
    if cob:
        wsr.append([])
        wsr.append(["Cobertura: con puntaje", cob.get("con_puntaje", 0), "/", cob.get("total_propuestas", 0)])
        wsr.append(["Sin puntaje", cob.get("sin_puntaje", 0)])
    for c in wsr["A"]:
        c.font = Font(bold=True)
    # Hoja por grupo
    for g in rk.get("grupos", []):
        title = (g.get("grupo") or "Sin grupo")[:31] or "Grupo"
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for idx, p in enumerate(g.get("items", []), 1):
            ws.append([
                idx, p.get("codigo"), p.get("nombre"), p.get("organizacion"),
                p.get("municipio"), p.get("subregion") or g.get("grupo"),
                p.get("puntaje_total"), p.get("bono_priorizacion") or 0,
                p.get("puntaje_final") or p.get("puntaje_total"),
                p.get("puntaje_diferencial") or 0,
                p.get("fuente") or "—", p.get("estado") or "—",
                p.get("cupo") or "", p.get("regla_desempate") or "",
            ])
        # ancho
        for col_idx, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = max(12, len(h) + 2)
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ranking_{rid[:8]}.xlsx"'}
    )


# ============================================================
# CUPOS DE GANADORES POR GRUPO (subregión, línea, etc.)
# ============================================================
@router.get("/convocatorias/{cid}/cupos-ganadores")
async def get_cupos(cid: str, user: dict = Depends(get_current_user)):
    db = get_db()
    conv = await db.convocatorias.find_one({"id": cid}, {"_id": 0, "configuracion": 1, "codigo": 1})
    if not conv:
        raise HTTPException(404, "Convocatoria no encontrada")
    cupos = (conv.get("configuracion") or {}).get("cupos_ganadores") or {}
    # Si es INC2026 y no hay cupos configurados, devolver los defaults como sugerencia
    suggested = None
    if (conv.get("codigo", "").upper() == "INC2026") and not cupos.get("subregion"):
        suggested = {"subregion": DEFAULT_CUPOS_INC2026_SUBREGION}
    return {"cupos": cupos, "suggested": suggested}


@router.patch("/convocatorias/{cid}/cupos-ganadores")
async def set_cupos(cid: str, payload: dict = Body(...),
                     user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    """payload: { agrupacion: 'subregion', cupos: { 'Urabá': 14, ... } } o { reset: true }"""
    db = get_db()
    conv = await db.convocatorias.find_one({"id": cid})
    if not conv:
        raise HTTPException(404, "Convocatoria no encontrada")
    config = conv.get("configuracion") or {}
    cupos_all = config.get("cupos_ganadores") or {}
    if payload.get("reset"):
        cupos_all = {}
    elif payload.get("seed_inc2026"):
        cupos_all["subregion"] = DEFAULT_CUPOS_INC2026_SUBREGION
    else:
        agrup = payload.get("agrupacion")
        cupos = payload.get("cupos") or {}
        if not agrup:
            raise HTTPException(400, "Falta 'agrupacion'")
        # Sanitizar a ints
        cupos_all[agrup] = {k: int(v) for k, v in cupos.items() if isinstance(v, (int, float)) and v >= 0}
    config["cupos_ganadores"] = cupos_all
    await db.convocatorias.update_one({"id": cid}, {"$set": {"configuracion": config}})
    await audit(user, "update", "cupos_ganadores", cid)
    return {"ok": True, "cupos": cupos_all}


# ==================== DASHBOARD ====================
@router.get("/dashboard")
async def dashboard(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    total = await db.propuestas.count_documents({"convocatoria_id": convocatoria_id})
    habilitadas = await db.propuestas.count_documents({"convocatoria_id": convocatoria_id, "estado": "Habilitada"})
    no_hab = await db.propuestas.count_documents({"convocatoria_id": convocatoria_id, "estado": "No habilitada"})
    asignadas = await db.propuestas.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$in": ["Asignada", "En evaluación individual", "En evaluación colectiva", "Consolidada", "Rankeada"]}})

    eval_ind_pend = await db.evaluaciones_individuales.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$in": ["Borrador", "Iniciada", "En edición"]}})
    eval_ind_fin = await db.evaluaciones_individuales.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$in": ["Finalizada", "Firmada"]}})
    eval_col_abr = await db.evaluaciones_colectivas.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$in": ["Abierta", "En proceso"]}})
    eval_col_cer = await db.evaluaciones_colectivas.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$in": ["Cerrada", "Firmada"]}})
    jurados_act = await db.jurados.count_documents({"convocatoria_id": convocatoria_id, "estado": "Activo"})
    ternas_act = await db.ternas.count_documents({"convocatoria_id": convocatoria_id, "estado": {"$ne": "Inactivo"}})

    # Avance por subregión (basado en datos.subregion)
    by_sub = {}
    cur = db.propuestas.find({"convocatoria_id": convocatoria_id}, {"datos.subregion": 1, "_id": 0})
    async for p in cur:
        s = (p.get("datos") or {}).get("subregion") or "Sin subregión"
        by_sub[s] = by_sub.get(s, 0) + 1
    avance_subregion = [{"subregion": k, "total": v} for k, v in sorted(by_sub.items(), key=lambda x: -x[1])]

    return {
        "total_propuestas": total,
        "habilitadas": habilitadas,
        "no_habilitadas": no_hab,
        "asignadas": asignadas,
        "evaluaciones_individuales_pendientes": eval_ind_pend,
        "evaluaciones_individuales_finalizadas": eval_ind_fin,
        "evaluaciones_colectivas_abiertas": eval_col_abr,
        "evaluaciones_colectivas_cerradas": eval_col_cer,
        "jurados_activos": jurados_act,
        "ternas_activas": ternas_act,
        "avance_subregion": avance_subregion,
    }


# ==================== REPORTES ====================
@router.get("/reportes/avance-jurado")
async def reporte_avance_jurado(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    jurados = await db.jurados.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(2000)
    out = []
    for j in jurados:
        asign = await db.asignaciones.count_documents({"convocatoria_id": convocatoria_id, "jurado_id": j["id"], "tipo_evaluacion": "individual"})
        iniciadas = await db.evaluaciones_individuales.count_documents({"jurado_id": j["id"], "estado": {"$in": ["Iniciada", "En edición"]}})
        fin = await db.evaluaciones_individuales.count_documents({"jurado_id": j["id"], "estado": {"$in": ["Finalizada", "Firmada"]}})
        firm = await db.evaluaciones_individuales.count_documents({"jurado_id": j["id"], "estado": "Firmada"})
        pend = max(asign - fin, 0)
        out.append({
            "jurado": j["nombre"], "correo": j["email"],
            "propuestas_asignadas": asign, "evaluaciones_iniciadas": iniciadas,
            "evaluaciones_finalizadas": fin, "evaluaciones_firmadas": firm,
            "pendientes": pend, "porcentaje_avance": round((fin / asign * 100) if asign else 0, 1)
        })
    return out


@router.get("/reportes/avance-terna")
async def reporte_avance_terna(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    ternas = await db.ternas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(500)
    # Resolver nombres de jurados desde la colección (no confiar en cache de la terna)
    juradoMap = {j["id"]: j.get("nombre") or j.get("email") or j["id"]
                 async for j in db.jurados.find({"convocatoria_id": convocatoria_id})}
    out = []
    for t in ternas:
        prop_count = await db.asignaciones.count_documents({"terna_id": t["id"]})
        col_abr = await db.evaluaciones_colectivas.count_documents({"terna_id": t["id"], "estado": {"$in": ["Abierta", "Reabierta", "En proceso"]}})
        col_cer = await db.evaluaciones_colectivas.count_documents({"terna_id": t["id"], "estado": {"$in": ["Cerrada", "Firmada"]}})
        integrantes = ", ".join([juradoMap.get(i.get("jurado_id"), i.get("nombre") or i.get("jurado_id") or "?")
                                  for i in t.get("integrantes", [])])
        out.append({
            "codigo": t["codigo"], "nombre": t["nombre"], "integrantes": integrantes,
            "propuestas_asignadas": prop_count, "colectivas_abiertas": col_abr,
            "colectivas_cerradas": col_cer,
            "porcentaje_avance": round((col_cer / prop_count * 100) if prop_count else 0, 1)
        })
    return out


@router.get("/reportes/consolidado-individual")
async def reporte_consolidado_individual(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Sábana completa de evaluación individual: una fila por (propuesta, jurado)
    con criterios, puntajes y observaciones. Excluye V2 colectivas.
    """
    db = get_db()
    evals = await db.evaluaciones_individuales.find({
        "convocatoria_id": convocatoria_id, "etapa": {"$ne": "colectiva"},
    }, {"_id": 0}).to_list(20000)
    propuestas = {p["id"]: p async for p in db.propuestas.find({"convocatoria_id": convocatoria_id})}
    jurados = {j["id"]: j async for j in db.jurados.find({"convocatoria_id": convocatoria_id})}
    criterios = await db.criterios.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("orden", 1).to_list(200)
    oficiales = [c for c in criterios if c.get("oficial") and not c.get("diferencial")]
    desempate = [c for c in criterios if c.get("diferencial")]
    out = []
    for ev in evals:
        p = propuestas.get(ev["propuesta_id"], {})
        j = jurados.get(ev["jurado_id"], {})
        row = {
            "propuesta_codigo": p.get("codigo"),
            "propuesta_nombre": p.get("nombre"),
            "organizacion": p.get("organizacion") or (p.get("datos") or {}).get("nombre_organizacion"),
            "subregion": (p.get("datos") or {}).get("subregion"),
            "jurado": j.get("nombre"),
            "jurado_email": j.get("email"),
            "estado": ev.get("estado"),
            "puntaje_total_oficial": ev.get("puntaje_total"),
            "puntaje_total_priorizacion": ev.get("puntaje_diferencial_total"),
        }
        puntajes = ev.get("puntajes") or {}
        observaciones = ev.get("observaciones") or {}
        # Columnas dinámicas por criterio OFICIAL (puntaje + observación)
        for c in oficiales:
            base = f"OF · {c.get('nombre','')}"
            row[f"{base} (puntaje)"] = puntajes.get(c["id"])
            row[f"{base} (obs.)"] = observaciones.get(c["id"], "")
        # Columnas dinámicas por criterio DIFERENCIAL/PRIORIZACIÓN/DESEMPATE
        for c in desempate:
            base = f"DIF · {c.get('nombre','')}"
            row[f"{base} (puntaje)"] = puntajes.get(c["id"])
            row[f"{base} (obs.)"] = observaciones.get(c["id"], "")
        row["observacion_final"] = ev.get("observacion_final")
        row["fecha_finalizacion"] = ev.get("fecha_finalizacion")
        row["reaperturas"] = ev.get("reaperturas", 0)
        out.append(row)
    return out


@router.get("/reportes/consolidado-colectiva")
async def reporte_consolidado_colectiva(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Reporte consolidado de evaluaciones colectivas por terna.

    Estructura: una fila por (propuesta, terna). Incluye:
      - Datos de la propuesta (código, nombre, organización, NIT, subregión, municipio).
      - Identificación de la terna y sus integrantes.
      - Puntaje TOTAL emitido por cada uno de los 3 jurados (suma de criterios oficiales de su V2).
      - Puntaje TOTAL promedio (validado contra el cierre automático de la colectiva).
      - Estado y fechas.

    No se incluyen observaciones por criterio porque en la evaluación colectiva
    el promedio se calcula numéricamente y no existe una observación consensuada por criterio.
    """
    db = get_db()
    evals = await db.evaluaciones_colectivas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(20000)
    propuestas = {p["id"]: p async for p in db.propuestas.find({"convocatoria_id": convocatoria_id})}
    ternas = {t["id"]: t async for t in db.ternas.find({"convocatoria_id": convocatoria_id})}
    juradoMap = {j["id"]: j async for j in db.jurados.find({"convocatoria_id": convocatoria_id})}
    criterios = await db.criterios.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("orden", 1).to_list(200)
    crit_oficiales_ids = [c["id"] for c in criterios if c.get("oficial") and not c.get("diferencial")]

    # Pre-cargar todas las V2 de la convocatoria (etapa=colectiva, finalizadas/firmadas)
    v2s = await db.evaluaciones_individuales.find({
        "convocatoria_id": convocatoria_id, "etapa": "colectiva",
        "estado": {"$in": ["Finalizada", "Firmada"]},
    }, {"_id": 0}).to_list(20000)
    # Map: (propuesta_id, jurado_id) -> total_oficial
    v2_total = {}
    for v in v2s:
        pj_total = v.get("puntaje_total")
        if pj_total is None:
            pj_total = sum(float((v.get("puntajes") or {}).get(cid, 0)) for cid in crit_oficiales_ids)
        v2_total[(v["propuesta_id"], v["jurado_id"])] = round(float(pj_total), 2)

    out = []
    for ev in evals:
        p = propuestas.get(ev.get("propuesta_id"), {})
        t = ternas.get(ev.get("terna_id"), {})
        d = p.get("datos") or {}
        integrantes_ids = [i.get("jurado_id") for i in (t.get("integrantes") or []) if i.get("jurado_id")]
        # Resolver nombres
        nombres = [juradoMap.get(jid, {}).get("nombre", "—") for jid in integrantes_ids]
        # Puntajes por jurado (en orden de la terna)
        puntajes_jur = [v2_total.get((ev.get("propuesta_id"), jid)) for jid in integrantes_ids]
        # Asegurar 3 columnas
        while len(puntajes_jur) < 3:
            puntajes_jur.append(None)
            nombres.append("—")
        # Promedio calculado
        vals_validos = [v for v in puntajes_jur if v is not None]
        promedio_calc = round(sum(vals_validos) / len(vals_validos), 2) if vals_validos else None
        puntaje_oficial = ev.get("puntaje_criterios") or ev.get("puntaje_total") or ev.get("puntaje_consensuado")
        row = {
            "propuesta_codigo": p.get("codigo"),
            "propuesta_nombre": p.get("nombre"),
            "organizacion": p.get("organizacion") or d.get("nombre_organizacion"),
            "nit": p.get("nit") or d.get("nit") or d.get("NIT") or d.get("numero_documento"),
            "municipio": d.get("municipio"),
            "subregion": d.get("subregion") or p.get("subregion"),
            "terna_codigo": t.get("codigo"),
            "terna_nombre": t.get("nombre"),
            "jurado_1": nombres[0],
            "puntaje_jurado_1": puntajes_jur[0],
            "jurado_2": nombres[1],
            "puntaje_jurado_2": puntajes_jur[1],
            "jurado_3": nombres[2],
            "puntaje_jurado_3": puntajes_jur[2],
            "puntaje_total_promedio_calculado": promedio_calc,
            "puntaje_total_oficial": round(float(puntaje_oficial), 2) if puntaje_oficial is not None else None,
            "bono_priorizacion": ev.get("bono_priorizacion") or 0,
            "puntaje_final_con_bono": ev.get("puntaje_final"),
            "puntaje_diferencial_total": ev.get("puntaje_diferencial_total"),
            "estado": ev.get("estado"),
            "fecha_cierre": ev.get("fecha_cierre") or ev.get("fecha_finalizacion"),
        }
        out.append(row)
    return out


@router.get("/reportes/consolidado-colectiva-detallado")
async def reporte_consolidado_colectiva_detallado(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Sábana completa de la EVALUACIÓN COLECTIVA: una fila por (propuesta, jurado de la terna)
    con TODOS los criterios desglosados (oficiales con observación + diferenciales/priorización/desempate).
    Es el equivalente del consolidado-individual pero usando las V2 de la etapa colectiva.
    """
    db = get_db()
    v2s = await db.evaluaciones_individuales.find({
        "convocatoria_id": convocatoria_id, "etapa": "colectiva",
    }, {"_id": 0}).to_list(20000)
    propuestas = {p["id"]: p async for p in db.propuestas.find({"convocatoria_id": convocatoria_id})}
    jurados = {j["id"]: j async for j in db.jurados.find({"convocatoria_id": convocatoria_id})}
    ternas_list = await db.ternas.find({"convocatoria_id": convocatoria_id}).to_list(500)
    jurado_terna = {}
    for t in ternas_list:
        for i in (t.get("integrantes") or []):
            if i.get("jurado_id"):
                jurado_terna.setdefault(i["jurado_id"], []).append(t.get("codigo"))
    criterios = await db.criterios.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("orden", 1).to_list(200)
    oficiales = [c for c in criterios if c.get("oficial") and not c.get("diferencial")]
    desempate = [c for c in criterios if c.get("diferencial")]
    out = []
    for ev in v2s:
        p = propuestas.get(ev["propuesta_id"], {})
        j = jurados.get(ev["jurado_id"], {})
        d = p.get("datos") or {}
        row = {
            "propuesta_codigo": p.get("codigo"),
            "propuesta_nombre": p.get("nombre"),
            "organizacion": p.get("organizacion") or d.get("nombre_organizacion"),
            "nit": p.get("nit") or d.get("nit") or d.get("NIT") or d.get("numero_documento"),
            "municipio": d.get("municipio"),
            "subregion": d.get("subregion") or p.get("subregion"),
            "terna": ", ".join(jurado_terna.get(ev["jurado_id"], [])),
            "jurado": j.get("nombre"),
            "jurado_email": j.get("email"),
            "estado": ev.get("estado"),
            "puntaje_total_oficial": ev.get("puntaje_total"),
            "puntaje_total_priorizacion": ev.get("puntaje_diferencial_total"),
        }
        puntajes = ev.get("puntajes") or {}
        observaciones = ev.get("observaciones") or {}
        for c in oficiales:
            base = f"OF · {c.get('nombre','')}"
            row[f"{base} (puntaje)"] = puntajes.get(c["id"])
            row[f"{base} (obs.)"] = observaciones.get(c["id"], "")
        for c in desempate:
            base = f"DIF · {c.get('nombre','')}"
            row[f"{base} (puntaje)"] = puntajes.get(c["id"])
            row[f"{base} (obs.)"] = observaciones.get(c["id"], "")
        row["observacion_final"] = ev.get("observacion_final")
        row["fecha_finalizacion"] = ev.get("fecha_finalizacion")
        out.append(row)
    return out


@router.get("/reportes/auditoria")
async def reporte_auditoria(limit: int = 500, entidad: Optional[str] = None,
                            user: dict = Depends(require_roles("admin_general", "auditor"))):
    db = get_db()
    q = {}
    if entidad: q["entidad"] = entidad
    items = await db.auditoria.find(q, {"_id": 0}).sort("fecha", -1).to_list(limit)
    return items


# ==================== EXPORT EXCEL ====================
@router.get("/reportes/export-excel")
async def export_excel(reporte: str, convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    wb = Workbook(); ws = wb.active; ws.title = reporte[:30]
    if reporte == "avance-jurado":
        data = await reporte_avance_jurado(convocatoria_id, user)
        headers = ["jurado", "correo", "propuestas_asignadas", "evaluaciones_iniciadas",
                   "evaluaciones_finalizadas", "evaluaciones_firmadas", "pendientes", "porcentaje_avance"]
    elif reporte == "avance-terna":
        data = await reporte_avance_terna(convocatoria_id, user)
        headers = ["codigo", "nombre", "integrantes", "propuestas_asignadas",
                   "colectivas_abiertas", "colectivas_cerradas", "porcentaje_avance"]
    elif reporte == "consolidado-individual":
        data = await reporte_consolidado_individual(convocatoria_id, user)
        headers = list(data[0].keys()) if data else [
            "propuesta_codigo", "propuesta_nombre", "organizacion", "jurado", "estado",
            "puntaje_total_oficial", "puntaje_total_priorizacion", "observacion_final", "fecha_finalizacion",
        ]
    elif reporte == "consolidado-colectiva":
        data = await reporte_consolidado_colectiva(convocatoria_id, user)
        headers = list(data[0].keys()) if data else [
            "propuesta_codigo", "propuesta_nombre", "organizacion", "terna_codigo", "terna_nombre", "estado",
            "puntaje_total_oficial", "puntaje_total_priorizacion", "observacion_final", "fecha_finalizacion",
        ]
    elif reporte == "consolidado-colectiva-detallado":
        data = await reporte_consolidado_colectiva_detallado(convocatoria_id, user)
        headers = list(data[0].keys()) if data else ["propuesta_codigo", "jurado"]
    else:
        raise HTTPException(status_code=400, detail="Reporte no soportado")
    ws.append(headers)
    # Estilo header
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10.5)
        cell.fill = PatternFill("solid", fgColor="14776A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for row in data:
        ws.append([row.get(h, "") for h in headers])
    # Anchos
    for i, h in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if "obs" in h.lower() or "observac" in h.lower():
            ws.column_dimensions[letter].width = 50
        elif "nombre" in h.lower() or "organizacion" in h.lower():
            ws.column_dimensions[letter].width = 32
        elif "puntaje" in h.lower() or "fecha" in h.lower():
            ws.column_dimensions[letter].width = 14
        else:
            ws.column_dimensions[letter].width = max(14, min(28, len(h) + 4))
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={reporte}.xlsx"})


# ==================== ACTAS PDF ====================
def _build_acta(title: str, conv: dict, body_rows: list, observacion: str = "", firmantes: list = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, spaceAfter=12, textColor=rl_colors.HexColor("#09090B"))
    subtitle = ParagraphStyle("sub", parent=styles["Heading3"], fontSize=11, textColor=rl_colors.HexColor("#52525B"))
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=14)
    elements = []
    entidad = (conv.get("entidades") or [{}])[0]
    elements.append(Paragraph(f"<b>KRINOS</b> — {entidad.get('nombre','')}", subtitle))
    elements.append(Paragraph(conv.get("nombre", ""), subtitle))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Código convocatoria: <b>{conv.get('codigo','')}</b> · Vigencia {conv.get('vigencia','')}", body))
    elements.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body))
    elements.append(Spacer(1, 16))

    if body_rows:
        tbl = Table(body_rows, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#059669")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 12))

    if observacion:
        elements.append(Paragraph("<b>Observación consolidada</b>", subtitle))
        elements.append(Paragraph(observacion, body))
        elements.append(Spacer(1, 12))

    if firmantes:
        elements.append(Spacer(1, 24))
        elements.append(Paragraph("<b>Firmantes</b>", subtitle))
        sig_rows = [["Nombre", "Rol", "Fecha"]] + firmantes
        sig_tbl = Table(sig_rows, hAlign="LEFT")
        sig_tbl.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, rl_colors.HexColor("#52525B")),
        ]))
        elements.append(sig_tbl)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


@router.get("/actas/individual/{eval_id}")
async def acta_individual(eval_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    ev = await db.evaluaciones_individuales.find_one({"id": eval_id})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    conv = await db.convocatorias.find_one({"id": ev["convocatoria_id"]})
    prop = await db.propuestas.find_one({"id": ev["propuesta_id"]})
    jur = await db.jurados.find_one({"id": ev["jurado_id"]})
    crits = await db.criterios.find({"convocatoria_id": ev["convocatoria_id"]}).sort("orden", 1).to_list(50)
    rows = [["Criterio", "Puntaje", "Observación"]]
    for c in crits:
        rows.append([c["nombre"], str(ev.get("puntajes", {}).get(c["id"], "—")), ev.get("observaciones", {}).get(c["id"], "")[:200]])
    rows.append(["TOTAL OFICIAL", str(ev.get("puntaje_total", 0)), ""])
    rows.append(["TOTAL DIFERENCIAL", str(ev.get("puntaje_diferencial_total", 0)), ""])
    firmantes = [[jur["nombre"] if jur else "—", "Jurado evaluador", ev.get("fecha_firma") or ev.get("fecha_finalizacion") or ""]]
    pdf = _build_acta(
        f"Acta de Evaluación Individual — {prop['codigo']}",
        conv or {}, rows,
        observacion=ev.get("observacion_final", ""), firmantes=firmantes
    )
    await audit(user, "generate_acta", "actas", eval_id, detalle="individual")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="acta_individual_{prop["codigo"]}.pdf"'})


@router.get("/actas/colectiva/{eval_id}")
async def acta_colectiva(eval_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    ev = await db.evaluaciones_colectivas.find_one({"id": eval_id})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluación colectiva no encontrada")
    conv = await db.convocatorias.find_one({"id": ev["convocatoria_id"]})
    prop = await db.propuestas.find_one({"id": ev["propuesta_id"]})
    terna = await db.ternas.find_one({"id": ev["terna_id"]})
    crits = await db.criterios.find({"convocatoria_id": ev["convocatoria_id"]}).sort("orden", 1).to_list(50)
    rows = [["Criterio", "Puntaje colectivo"]]
    for c in crits:
        rows.append([c["nombre"], str(ev.get("puntajes", {}).get(c["id"], "—"))])
    rows.append(["PUNTAJE FINAL COLECTIVO", str(ev.get("puntaje_final", 0))])
    firmantes = []
    for integ in (terna or {}).get("integrantes", []):
        firmantes.append([integ.get("nombre", integ.get("jurado_id", "—")), integ.get("rol", "Integrante"), ev.get("fecha_cierre", "")])
    pdf = _build_acta(
        f"Acta de Evaluación Colectiva — {prop['codigo']} · {terna['codigo']}" if prop and terna else "Acta Colectiva",
        conv or {}, rows,
        observacion=ev.get("observacion_consolidada", ""), firmantes=firmantes
    )
    await audit(user, "generate_acta", "actas", eval_id, detalle="colectiva")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="acta_colectiva_{prop["codigo"] if prop else eval_id}.pdf"'})


@router.get("/actas/ranking/{ranking_id}")
async def acta_ranking(ranking_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    rk = await db.rankings.find_one({"id": ranking_id})
    if not rk:
        raise HTTPException(status_code=404, detail="Ranking no encontrado")
    conv = await db.convocatorias.find_one({"id": rk["convocatoria_id"]})
    rows = [["Puesto", "Grupo", "Código", "Propuesta", "Organización", "Puntaje", "Desempate aplicado"]]
    for g in rk["grupos"]:
        for it in g["items"]:
            rows.append([str(it["puesto"]), g["grupo"], it["codigo"], it["nombre"][:50],
                         (it["organizacion"] or "")[:30], str(it["puntaje_total"]),
                         it.get("desempate_regla", "—")])
    pdf = _build_acta(
        f"Acta de Ranking — Agrupado por {rk['agrupacion']}",
        conv or {}, rows,
        observacion=f"Modo: {rk['modo']} · Fecha generación: {rk['fecha_generacion']}",
    )
    await audit(user, "generate_acta", "actas", ranking_id, detalle="ranking")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="acta_ranking_{rk["id"][:8]}.pdf"'})
