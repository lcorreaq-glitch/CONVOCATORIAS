"""Iter 11 - Backend tests for: campos aplica_a, upload/file, ai/mejorar-texto,
jurados parametrizable, jurados/me, jurados-template dynamic.
"""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/") or "https://convocatoria-hub-2.preview.emergentagent.com"
ADMIN_USER = "lcorreaq"
ADMIN_PASS = "Chocolate2026!"
CONV_CODE = "INC2026"


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=20)
    assert r.status_code == 200, f"Login admin failed: {r.status_code} {r.text}"
    data = r.json()
    tok = data.get("token") or data.get("access_token")
    assert tok, f"No token in response: {data}"
    return tok


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def CONV_ID(admin_headers):
    """Resolve INC2026 codigo -> internal UUID."""
    r = requests.get(f"{BASE_URL}/api/convocatorias", headers=admin_headers, timeout=15)
    assert r.status_code == 200, r.text
    for c in r.json():
        if c.get("codigo") == CONV_CODE:
            return c["id"]
    pytest.skip(f"Convocatoria {CONV_CODE} not found")


@pytest.fixture(scope="module")
def jurado_token():
    """Login as a seeded jurado (try several seed emails)."""
    candidates = [
        "yicell.gonzalez@fodc.org.co",
        "laura.arroyave@fodc.org.co",
        "ana.perez@krinos.gov.co",
    ]
    for u in candidates:
        r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": u, "password": "Jurado2026!"}, timeout=20)
        if r.status_code == 200:
            return r.json().get("token") or r.json().get("access_token")
    pytest.skip("No seeded jurado login worked")


# --------- Campos aplica_a ---------

class TestCamposAplicaA:
    def test_campos_jurado_returns_only_jurado(self, admin_headers, CONV_ID):
        r = requests.get(f"{BASE_URL}/api/campos",
                         params={"convocatoria_id": CONV_ID, "aplica_a": "jurado"},
                         headers=admin_headers, timeout=15)
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        # Spec says 8 jurado campos
        assert len(items) == 8, f"Expected 8 jurado campos, got {len(items)}: {[c.get('nombre_interno') for c in items]}"
        nombres = {c["nombre_interno"] for c in items}
        expected = {"nombre", "cedula", "organizacion", "subregiones", "telefono", "email", "perfil", "hoja_vida"}
        assert expected.issubset(nombres), f"Missing: {expected - nombres}"
        # All must have aplica_a == jurado
        for c in items:
            assert c.get("aplica_a") == "jurado", f"Campo {c.get('nombre_interno')} has aplica_a={c.get('aplica_a')}"

    def test_campos_propuesta_returns_only_propuesta(self, admin_headers, CONV_ID):
        r = requests.get(f"{BASE_URL}/api/campos",
                         params={"convocatoria_id": CONV_ID, "aplica_a": "propuesta"},
                         headers=admin_headers, timeout=15)
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        # Spec says 16 propuesta campos
        assert len(items) == 16, f"Expected 16 propuesta campos, got {len(items)}"
        for c in items:
            # null/missing aplica_a = propuesta (backwards compat)
            aa = c.get("aplica_a")
            assert aa in (None, "propuesta"), f"Campo {c.get('nombre_interno')} aplica_a={aa}"


# --------- Upload file ---------

class TestUploadFile:
    def test_upload_pdf(self, admin_headers):
        # Minimal valid-ish PDF
        pdf_bytes = b"%PDF-1.4\n1 0 obj<<>>endobj\ntrailer<<>>\n%%EOF\n"
        files = {"file": ("hoja_vida.pdf", io.BytesIO(pdf_bytes), "application/pdf")}
        r = requests.post(f"{BASE_URL}/api/upload/file", files=files, headers=admin_headers, timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "data_url" in data and data["data_url"].startswith("data:application/pdf;base64,")
        assert data["filename"] == "hoja_vida.pdf"
        assert data["size"] == len(pdf_bytes)
        assert data["content_type"] == "application/pdf"

    def test_upload_rejects_bad_type(self, admin_headers):
        files = {"file": ("evil.exe", io.BytesIO(b"MZ\x00"), "application/x-msdownload")}
        r = requests.post(f"{BASE_URL}/api/upload/file", files=files, headers=admin_headers, timeout=15)
        assert r.status_code == 400


# --------- AI mejorar-texto ---------

class TestAIMejorarTexto:
    def test_mejorar_texto_perfil_jurado(self, admin_headers):
        r = requests.post(f"{BASE_URL}/api/ai/mejorar-texto",
                          json={"texto": "Soy un evaluador con experiencia en cultura y emprendimientos comunitarios.",
                                "contexto": "perfil_jurado"},
                          headers=admin_headers, timeout=60)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "texto_mejorado" in data
        assert "texto_original" in data
        assert len(data["texto_mejorado"]) > 10
        assert data["texto_original"].startswith("Soy un evaluador")

    def test_mejorar_texto_empty_rejected(self, admin_headers):
        r = requests.post(f"{BASE_URL}/api/ai/mejorar-texto",
                          json={"texto": "ab", "contexto": "perfil_jurado"},
                          headers=admin_headers, timeout=15)
        assert r.status_code == 400


# --------- Jurados list ---------

class TestJurados:
    def test_list_jurados_inc2026(self, admin_headers, CONV_ID):
        r = requests.get(f"{BASE_URL}/api/jurados",
                         params={"convocatoria_id": CONV_ID},
                         headers=admin_headers, timeout=20)
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        assert len(items) == 29, f"Expected 29 seeded jurados, got {len(items)}"
        # Spot-check structure
        sample = items[0]
        assert "subregiones" in sample
        assert isinstance(sample["subregiones"], list)
        assert "datos" in sample
        assert isinstance(sample["datos"], dict)
        # At least some have cedula or organizacion in datos
        with_cedula = [j for j in items if j.get("datos", {}).get("cedula")]
        with_org = [j for j in items if j.get("datos", {}).get("organizacion")]
        assert len(with_cedula) > 0, "Ningún jurado tiene datos.cedula"
        assert len(with_org) > 0, "Ningún jurado tiene datos.organizacion"

    def test_jurados_template_xlsx(self, admin_headers, CONV_ID):
        r = requests.get(f"{BASE_URL}/api/jurados-template",
                         params={"convocatoria_id": CONV_ID},
                         headers=admin_headers, timeout=20)
        assert r.status_code == 200
        ctype = r.headers.get("content-type", "")
        assert "spreadsheet" in ctype, ctype
        # Parse and check headers
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0])
        # Expect dynamic columns from jurado campos
        for col in ["nombre", "email", "telefono", "perfil"]:
            assert col in headers, f"Missing column {col} in template. Got: {headers}"


# --------- Jurados /me ---------

class TestJuradosMe:
    def test_me_admin_returns_404(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/jurados/me", headers=admin_headers, timeout=15)
        assert r.status_code == 404, f"Admin (no jurado_id) should get 404, got {r.status_code}"

    def test_me_jurado_returns_record(self, jurado_token):
        h = {"Authorization": f"Bearer {jurado_token}"}
        r = requests.get(f"{BASE_URL}/api/jurados/me", headers=h, timeout=15)
        assert r.status_code == 200, r.text
        jur = r.json()
        assert "id" in jur
        assert "nombre" in jur
        assert "email" in jur

    def test_patch_me_ignores_unsafe_fields(self, jurado_token):
        h = {"Authorization": f"Bearer {jurado_token}"}
        # Get current state
        before = requests.get(f"{BASE_URL}/api/jurados/me", headers=h, timeout=15).json()
        orig_nombre = before["nombre"]
        orig_email = before["email"]
        orig_subs = before.get("subregiones", [])

        # Try to patch unsafe fields + safe (telefono, perfil)
        payload = {
            "nombre": "HACKED",
            "email": "hacker@evil.com",
            "subregiones": ["FAKE"],
            "telefono": "3009998888",
            "perfil": "Perfil de prueba iter11 - safe update.",
        }
        r = requests.patch(f"{BASE_URL}/api/jurados/me", headers=h, json=payload, timeout=15)
        assert r.status_code == 200, r.text
        after = r.json()
        assert after["nombre"] == orig_nombre, "nombre fue modificado!"
        assert after["email"] == orig_email, "email fue modificado!"
        assert after.get("subregiones", []) == orig_subs, "subregiones fueron modificadas!"
        assert after["telefono"] == "3009998888"
        assert after["perfil"] == "Perfil de prueba iter11 - safe update."

        # Restore original telefono/perfil to avoid drift
        requests.patch(f"{BASE_URL}/api/jurados/me", headers=h, json={
            "telefono": before.get("telefono") or "",
            "perfil": before.get("perfil") or "",
        }, timeout=15)
