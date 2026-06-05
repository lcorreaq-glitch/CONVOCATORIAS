"""KRINOS - Data: propuestas, jurados, ternas, asignaciones + bulk Excel import/export."""

import uuid
import io
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook

from db import get_db, now_iso
from auth import get_current_user, require_roles, audit, hash_password

router = APIRouter(prefix="/api", tags=["data"])


# ==================== PROPUESTAS ====================
class PropuestaIn(BaseModel):
    convocatoria_id: str
    codigo: Optional[str] = None
    nombre: str
    organizacion: Optional[str] = ""
    datos: dict = Field(default_factory=dict)
    estado: str = "Registrada"


@router.get("/propuestas")
async def list_propuestas(
    convocatoria_id: str,
    estado: Optional[str] = None,
    subregion: Optional[str] = None,
    linea: Optional[str] = None,
    search: Optional[str] = None,
    filtros: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Lista propuestas con filtros dinámicos.

    `filtros` es un JSON con {nombre_interno_campo: valor} que se traduce a query
    sobre datos.<nombre_interno>. Soporta también si_no (true/false) y arrays
    (busca propuestas cuyo array contenga el valor).
    """
    import json as _json

    db = get_db()
    q = {"convocatoria_id": convocatoria_id}
    if estado:
        q["estado"] = estado
    if subregion:
        q["datos.subregion"] = subregion
    if linea:
        q["datos.linea"] = linea
    if filtros:
        try:
            extra = _json.loads(filtros)
            for k, v in (extra or {}).items():
                if v is None or v == "" or v == "__all__":
                    continue
                # Si es array (seleccion_multiple), buscar coincidencia
                if isinstance(v, list):
                    q[f"datos.{k}"] = {"$in": v}
                else:
                    q[f"datos.{k}"] = v
        except Exception:
            pass
    if search:
        q["$or"] = [
            {"nombre": {"$regex": search, "$options": "i"}},
            {"organizacion": {"$regex": search, "$options": "i"}},
            {"codigo": {"$regex": search, "$options": "i"}},
        ]
    items = await db.propuestas.find(q, {"_id": 0}).sort("codigo", 1).to_list(5000)
    return items


@router.get("/propuestas/{pid}")
async def get_propuesta(pid: str, user: dict = Depends(get_current_user)):
    db = get_db()
    item = await db.propuestas.find_one({"id": pid}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    return item


@router.post("/propuestas")
async def create_propuesta(payload: PropuestaIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    if not doc.get("codigo"):
        count = await db.propuestas.count_documents({"convocatoria_id": doc["convocatoria_id"]})
        doc["codigo"] = f"P-{count + 1:04d}"
    doc["created_at"] = now_iso()
    await db.propuestas.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "create", "propuestas", doc["id"], valor_nuevo={"codigo": doc["codigo"]})
    return doc


@router.patch("/propuestas/{pid}")
async def update_propuesta(pid: str, payload: dict, user: dict = Depends(require_roles("admin_general", "admin_convocatoria", "supervisor"))):
    db = get_db()
    payload.pop("id", None)
    await db.propuestas.update_one({"id": pid}, {"$set": payload})
    await audit(user, "update", "propuestas", pid, valor_nuevo=payload)
    return await db.propuestas.find_one({"id": pid}, {"_id": 0})


@router.get("/propuestas-template")
async def propuestas_template(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Descarga plantilla Excel para carga masiva de propuestas.
    La plantilla se construye dinámicamente a partir de los campos del formulario de propuesta:
    - Solo campos con `uso_propuesta != false`.
    - 2 filas de encabezado: etiqueta humana + nombre_interno (este último es el que se importa).
    - Marca obligatorios con asterisco.
    - Hoja "Instrucciones" con valores válidos por campo tipo catálogo.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    campos = await db.campos.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("orden", 1).to_list(500)
    # Filtrar solo los del formulario de propuesta (default True si no está definido)
    campos = [c for c in campos if c.get("uso_propuesta", True) is not False]
    catalogos = await db.catalogos.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(200)
    cat_by_id = {c["id"]: c for c in catalogos}
    cat_by_nombre = {c["nombre"]: c for c in catalogos}

    # Catálogo "Estados de Propuesta" (si existe se usa para validar/sugerir; si no, fallback estático)
    estados_cat = cat_by_nombre.get("Estados de Propuesta") or cat_by_nombre.get("Estados de propuesta")
    if estados_cat:
        estados_validos = [v.get("valor") for v in (estados_cat.get("valores") or []) if v.get("activo") is not False]
    else:
        estados_validos = ["Registrada", "En revisión", "Habilitada", "No habilitada", "Subsanación pendiente"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Propuestas"

    # Encabezados técnicos (lo que el importer lee). `codigo`, `nombre` y `estado` son top-level fijos.
    # La organización viene como campo dinámico (ej. `nombre_organizacion`) para no duplicar.
    headers_tech = ["codigo", "nombre"] + [c["nombre_interno"] for c in campos] + ["estado"]
    # Encabezados humanos
    headers_label = [
        "Código (opcional)",
        "Nombre de la propuesta *",
    ] + [((c.get("nombre_visible") or c["nombre_interno"]) + (" *" if c.get("obligatorio") else "")) for c in campos] + [
        "Estado (opcional)",
    ]

    # Fila 1: etiquetas humanas (visualmente clara para el usuario)
    ws.append(headers_label)
    label_font = Font(bold=True, color="FFFFFF", size=11)
    label_fill = PatternFill("solid", fgColor="14776A")
    for cell in ws[1]:
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Fila 2: nombres internos (importer lee de aquí)
    ws.append(headers_tech)
    intern_font = Font(italic=True, color="5E6878", size=9)
    intern_fill = PatternFill("solid", fgColor="F1F4F7")
    for cell in ws[2]:
        cell.font = intern_font
        cell.fill = intern_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Fila 3: ejemplo
    ws.append(["P-0001", "Mi propuesta ejemplo"] + ["" for _ in campos] + [estados_validos[0] if estados_validos else "Registrada"])

    # Ancho de columnas razonable
    for col_idx, h in enumerate(headers_label, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(16, min(38, len(h) + 4))
    ws.freeze_panes = "A3"

    # Mapeo de tipos: traduce los tipos reales del sistema a una descripción amigable.
    def _describe_field(c):
        tipo = c.get("tipo", "texto")
        # Tipos basados en catálogo
        if tipo in ("lista", "catalogo", "select"):
            cat_ref = c.get("catalogo_id") or c.get("catalogo")
            cat = cat_by_id.get(cat_ref) or cat_by_nombre.get(cat_ref or "")
            if cat:
                vals = [v.get("valor") for v in (cat.get("valores") or []) if v.get("activo") is not False]
                return "lista", f"Valor exacto de: {cat.get('nombre')}  →  Ver hoja 'Catálogos'"
            return tipo, "Texto libre"
        if tipo in ("multi_catalogo", "multi_select", "seleccion_multiple", "multiseleccion"):
            cat_ref = c.get("catalogo_id") or c.get("catalogo")
            cat = cat_by_id.get(cat_ref) or cat_by_nombre.get(cat_ref or "")
            if cat:
                return "multi-lista", f"Uno o varios separados por ;  de: {cat.get('nombre')}  →  Ver hoja 'Catálogos'"
            return tipo, "Texto libre (separar con ;)"
        if tipo in ("si_no", "boolean", "booleano"):
            return tipo, "sí | no  (también acepta true/false, 1/0)"
        if tipo == "fecha":
            return tipo, "Formato YYYY-MM-DD (ej. 2026-03-15)"
        if tipo == "hora":
            return tipo, "Formato HH:MM (24h, ej. 14:30)"
        if tipo in ("numero", "numero_entero", "numero_decimal", "entero", "decimal"):
            return tipo, "Número entero o decimal"
        if tipo in ("url", "enlace"):
            return tipo, "URL completa (https://...)"
        if tipo in ("texto_corto", "texto", "texto_largo", "textarea"):
            return tipo, "Texto libre"
        if tipo in ("email", "correo"):
            return tipo, "Correo electrónico válido"
        if tipo == "archivo":
            return tipo, "(Subir luego desde la UI, no se carga por Excel)"
        return tipo, "Texto libre"

    # Hoja "Instrucciones" con valores válidos de catálogos por campo
    inst = wb.create_sheet("Instrucciones")
    inst.append(["Campo (nombre interno)", "Etiqueta visible", "Tipo", "Obligatorio", "Valores aceptados / Catálogo"])
    for cell in inst[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F3F0")
    inst.append(["codigo", "Código", "texto", "No", "Opcional. Si vacío se autogenera (P-0001, P-0002...)"])
    inst.append(["nombre", "Nombre de la propuesta", "texto", "Sí", "Cualquier texto"])
    for c in campos:
        tipo_label, valores = _describe_field(c)
        inst.append(
            [
                c["nombre_interno"],
                c.get("nombre_visible") or c["nombre_interno"],
                tipo_label,
                "Sí" if c.get("obligatorio") else "No",
                valores,
            ]
        )
    # Fila final: estado
    inst.append(["estado", "Estado de la propuesta", "lista", "No",
                 ("Uno de: " + " | ".join(estados_validos)) if estados_validos else "Texto libre — por defecto 'Registrada'"])
    inst.column_dimensions["A"].width = 28
    inst.column_dimensions["B"].width = 32
    inst.column_dimensions["C"].width = 14
    inst.column_dimensions["D"].width = 12
    inst.column_dimensions["E"].width = 70

    # Hoja "Catálogos" — referencia rápida con todos los valores válidos
    cat_sheet = wb.create_sheet("Catálogos")
    cat_sheet.append(["Catálogo", "Valor", "Descripción / Padre"])
    for cell in cat_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14776A")
    # Solo catálogos referenciados por algún campo del formulario o por estado
    cat_ids_usados = set()
    for c in campos:
        ref = c.get("catalogo_id") or c.get("catalogo")
        if ref:
            cat_ids_usados.add(ref)
    if estados_cat:
        cat_ids_usados.add(estados_cat.get("id"))
    for cat in catalogos:
        if cat.get("id") not in cat_ids_usados and cat.get("nombre") not in cat_ids_usados:
            continue
        for v in (cat.get("valores") or []):
            if v.get("activo") is False:
                continue
            cat_sheet.append([cat.get("nombre"), v.get("valor"), v.get("descripcion") or v.get("padre_valor") or ""])
    cat_sheet.column_dimensions["A"].width = 28
    cat_sheet.column_dimensions["B"].width = 36
    cat_sheet.column_dimensions["C"].width = 40
    cat_sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_propuestas.xlsx"})


@router.post("/propuestas-import")
async def import_propuestas(convocatoria_id: str = Form(...), file: UploadFile = File(...), user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Archivo Excel inválido: {e}")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"creados": 0, "rechazados": 0, "errores": []}

    # Cargar catálogo "Estados de Propuesta" para validar la columna `estado`
    estados_cat = await db.catalogos.find_one(
        {"convocatoria_id": convocatoria_id, "nombre": {"$in": ["Estados de Propuesta", "Estados de propuesta"]}},
        {"_id": 0},
    )
    estados_validos_lower = set()
    if estados_cat:
        estados_validos_lower = {
            (v.get("valor") or "").strip().lower()
            for v in (estados_cat.get("valores") or [])
            if v.get("activo") is not False
        }

    # Detectar fila de encabezados técnicos (nombre_interno). Puede ser fila 1 (plantilla vieja)
    # o fila 2 (plantilla nueva con etiquetas humanas arriba).
    def _is_tech_header_row(r):
        if not r:
            return False
        first = (str(r[0] or "")).strip().lower()
        second = (str(r[1] or "")).strip().lower()
        return first == "codigo" and second == "nombre"

    header_row_idx = 0
    if not _is_tech_header_row(rows[0]) and len(rows) > 1 and _is_tech_header_row(rows[1]):
        header_row_idx = 1
    headers = [str(h).strip() if h else "" for h in rows[header_row_idx]]
    data_start = header_row_idx + 1
    created, errors = 0, []
    for idx, row in enumerate(rows[data_start:], start=data_start + 1):
        try:
            # Saltar filas vacías
            if not any(c for c in row if c not in (None, "", " ")):
                continue
            data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            if not data.get("nombre"):
                errors.append({"fila": idx, "error": "Falta nombre"})
                continue
            # Procesar estado (validar contra catálogo si existe)
            estado_raw = (data.get("estado") or "").strip() if isinstance(data.get("estado"), str) else (data.get("estado") or "")
            estado = "Registrada"
            if estado_raw:
                estado_str = str(estado_raw).strip()
                if estados_validos_lower:
                    if estado_str.lower() not in estados_validos_lower:
                        errors.append({"fila": idx, "error": f"Estado '{estado_str}' no es válido (ver hoja 'Catálogos')"})
                        continue
                    # Recuperar la forma canónica del valor
                    for v in (estados_cat.get("valores") or []):
                        if (v.get("valor") or "").strip().lower() == estado_str.lower():
                            estado = v.get("valor")
                            break
                else:
                    estado = estado_str
            datos = {
                k: v for k, v in data.items()
                if k not in ("codigo", "nombre", "organizacion", "estado") and v is not None
            }
            # Convert dates/times to string if needed
            for k, v in datos.items():
                if hasattr(v, "isoformat"):
                    datos[k] = v.isoformat()
            codigo = data.get("codigo") or f"P-{await db.propuestas.count_documents({'convocatoria_id': convocatoria_id}) + created + 1:04d}"
            # Detectar org dinámicamente: `organizacion`, `nombre_organizacion` o cualquier campo
            # del formulario que apunte a "organización". Permite que el formulario sea la fuente única.
            org = (data.get("organizacion") or data.get("nombre_organizacion")
                   or datos.get("nombre_organizacion") or datos.get("organizacion") or "")
            doc = {
                "id": str(uuid.uuid4()),
                "convocatoria_id": convocatoria_id,
                "codigo": str(codigo),
                "nombre": str(data["nombre"]),
                "organizacion": str(org),
                "datos": datos,
                "estado": estado,
                "created_at": now_iso(),
            }
            await db.propuestas.insert_one(doc)
            created += 1
        except Exception as e:
            errors.append({"fila": idx, "error": str(e)})
    await audit(user, "bulk_import", "propuestas", convocatoria_id, detalle=f"Creados {created}, errores {len(errors)}")
    return {"creados": created, "rechazados": len(errors), "errores": errors[:50]}


# ==================== JURADOS ====================
class JuradoIn(BaseModel):
    convocatoria_id: str
    nombre: str
    email: str
    telefono: Optional[str] = ""
    perfil: Optional[str] = ""
    especialidad: Optional[str] = ""
    linea_experiencia: Optional[str] = ""
    territorio: Optional[str] = ""
    subregiones: Optional[List[str]] = None
    disponibilidad: Optional[str] = "Disponible"
    estado: str = "Activo"
    crear_usuario: bool = True
    password: Optional[str] = None
    datos: Optional[dict] = None  # campos dinámicos definidos en Configuración
    foto_url: Optional[str] = None


@router.get("/jurados")
async def list_jurados(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    items = await db.jurados.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(2000)
    return items


@router.get("/jurados/me")
async def get_my_jurado(user: dict = Depends(get_current_user)):
    """Devuelve el registro de jurado vinculado al usuario actual."""
    db = get_db()
    if not user.get("jurado_id"):
        raise HTTPException(status_code=404, detail="Tu usuario no está vinculado a un jurado")
    jur = await db.jurados.find_one({"id": user["jurado_id"]}, {"_id": 0})
    if not jur:
        raise HTTPException(status_code=404, detail="Registro de jurado no encontrado")
    return jur


@router.patch("/jurados/me")
async def update_my_jurado(payload: dict, user: dict = Depends(get_current_user)):
    """Permite al jurado editar SOLO sus datos básicos (perfil, foto, datos dinámicos no críticos)."""
    db = get_db()
    if not user.get("jurado_id"):
        raise HTTPException(status_code=404, detail="Tu usuario no está vinculado a un jurado")
    # Campos que el propio jurado NO puede modificar (solo admin):
    SAFE_KEYS = {"telefono", "perfil", "especialidad", "linea_experiencia", "foto_url", "datos"}
    safe = {k: v for k, v in payload.items() if k in SAFE_KEYS}
    if not safe:
        return await db.jurados.find_one({"id": user["jurado_id"]}, {"_id": 0})
    await db.jurados.update_one({"id": user["jurado_id"]}, {"$set": safe})
    await audit(user, "self_update", "jurados", user["jurado_id"], valor_nuevo=safe)
    return await db.jurados.find_one({"id": user["jurado_id"]}, {"_id": 0})


@router.post("/jurados")
async def create_jurado(payload: JuradoIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    doc = payload.model_dump()
    pwd = doc.pop("password", None) or "Jurado2026!"
    crear_user = doc.pop("crear_usuario", True)
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    if doc.get("datos") is None:
        doc["datos"] = {}
    if doc.get("subregiones") is None:
        doc["subregiones"] = []
    await db.jurados.insert_one(doc)
    doc.pop("_id", None)

    credenciales = None
    if crear_user:
        username = doc["email"].lower()
        existing = await db.users.find_one({"$or": [{"username": username}, {"email": username}]})
        if not existing:
            await db.users.insert_one(
                {
                    "id": str(uuid.uuid4()),
                    "username": username,
                    "email": username,
                    "name": doc["nombre"],
                    "password_hash": hash_password(pwd),
                    "role": "jurado",
                    "active": True,
                    "convocatoria_roles": [{"convocatoria_id": doc["convocatoria_id"], "role": "jurado"}],
                    "jurado_id": doc["id"],
                    "created_at": now_iso(),
                }
            )
            # Devolver credenciales en claro UNA SOLA VEZ para envío por correo
            credenciales = {"username": username, "password": pwd, "rol": "jurado"}
    await audit(user, "create", "jurados", doc["id"], valor_nuevo={"nombre": doc["nombre"]})
    if credenciales:
        doc["credenciales"] = credenciales
    return doc


@router.patch("/jurados/{jid}")
async def update_jurado(jid: str, payload: dict, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    payload.pop("id", None)
    await db.jurados.update_one({"id": jid}, {"$set": payload})
    await audit(user, "update", "jurados", jid, valor_nuevo=payload)
    return await db.jurados.find_one({"id": jid}, {"_id": 0})


@router.get("/jurados-template")
async def jurados_template(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Plantilla dinámica de jurados, con mismo look & feel que la de propuestas:
    - 2 filas header (etiqueta humana + nombre interno) + fila de ejemplo.
    - Hoja Instrucciones con tipo/obligatorio/valores.
    - Hoja Catálogos con valores válidos (Subregiones, Estados de Jurado, ...).
    - Columna `estado` validada contra catálogo "Estados de Jurado" si existe.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    campos = await db.campos.find({"convocatoria_id": convocatoria_id, "aplica_a": "jurado"}, {"_id": 0}).sort("orden", 1).to_list(200)
    catalogos = await db.catalogos.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(200)
    cat_by_id = {c["id"]: c for c in catalogos}
    cat_by_nombre = {c["nombre"]: c for c in catalogos}

    subreg_cat = cat_by_nombre.get("Subregiones") or cat_by_nombre.get("subregiones")
    subreg_valores = [v.get("valor") for v in (subreg_cat.get("valores") or []) if v.get("activo") is not False] if subreg_cat else []

    estados_cat = cat_by_nombre.get("Estados de Jurado") or cat_by_nombre.get("Estados de jurado")
    estados_validos = [v.get("valor") for v in (estados_cat.get("valores") or []) if v.get("activo") is not False] if estados_cat else ["Activo", "Inactivo", "Suspendido"]

    # Filtrar campos a importar: excluir los de tipo archivo, los con rol_especial=firma/hoja_vida/foto,
    # y los que duplican las columnas base (nombre, email, telefono, subregiones, perfil).
    NO_IMPORT_ROLES = {"firma", "hoja_vida", "foto"}
    BASE_INTERNOS = {"nombre", "email", "telefono", "subregiones", "perfil"}
    campos_excel = [
        c for c in campos
        if c.get("tipo") != "archivo"
        and c.get("rol_especial") not in NO_IMPORT_ROLES
        and c.get("nombre_interno") not in BASE_INTERNOS
    ]

    base_specs = [
        ("nombre", "Nombre completo *", True, "Texto libre"),
        ("email", "Correo electrónico *", True, "Email válido (será su usuario de acceso)"),
        ("telefono", "Teléfono", False, "Ej. 3001234567"),
        ("subregiones", "Subregiones *", True, "Una o varias separadas por ; (ver hoja 'Catálogos')"),
        ("perfil", "Perfil profesional", False, "Texto libre (formación, experiencia)"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Jurados"

    headers_tech = [k for k, _, _, _ in base_specs] + [c["nombre_interno"] for c in campos_excel] + ["estado"]
    headers_label = [lbl for _, lbl, _, _ in base_specs] + [
        ((c.get("nombre_visible") or c["nombre_interno"]) + (" *" if c.get("obligatorio") else "")) for c in campos_excel
    ] + ["Estado (opcional)"]

    ws.append(headers_label)
    label_font = Font(bold=True, color="FFFFFF", size=11)
    label_fill = PatternFill("solid", fgColor="14776A")
    for cell in ws[1]:
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    ws.append(headers_tech)
    intern_font = Font(italic=True, color="5E6878", size=9)
    intern_fill = PatternFill("solid", fgColor="F1F4F7")
    for cell in ws[2]:
        cell.font = intern_font
        cell.fill = intern_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Fila de ejemplo
    ejemplo_subreg = subreg_valores[0] + ("; " + subreg_valores[1] if len(subreg_valores) > 1 else "") if subreg_valores else "Urabá; Norte"
    ejemplo = [
        "Ana María Pérez",
        "ana.perez@ejemplo.co",
        "3001234567",
        ejemplo_subreg,
        "Magíster en Desarrollo Comunitario con 10 años de experiencia.",
    ] + ["" for _ in campos_excel] + [estados_validos[0] if estados_validos else "Activo"]
    ws.append(ejemplo)

    for col_idx, h in enumerate(headers_label, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(18, min(38, len(h) + 4))
    ws.freeze_panes = "A3"

    # Mapeo de tipos amigable
    def _describe_field(c):
        tipo = c.get("tipo", "texto")
        if tipo in ("lista", "catalogo", "select"):
            cat_ref = c.get("catalogo_id") or c.get("catalogo")
            cat = cat_by_id.get(cat_ref) or cat_by_nombre.get(cat_ref or "")
            if cat:
                return "lista", f"Valor exacto de: {cat.get('nombre')}  →  Ver hoja 'Catálogos'"
            return tipo, "Texto libre"
        if tipo in ("multi_catalogo", "multi_select", "seleccion_multiple", "multiseleccion"):
            cat_ref = c.get("catalogo_id") or c.get("catalogo")
            cat = cat_by_id.get(cat_ref) or cat_by_nombre.get(cat_ref or "")
            if cat:
                return "multi-lista", f"Uno o varios separados por ;  de: {cat.get('nombre')}  →  Ver hoja 'Catálogos'"
            return tipo, "Texto libre (separar con ;)"
        if tipo in ("si_no", "boolean"):
            return tipo, "sí | no  (también acepta true/false, 1/0)"
        if tipo == "fecha":
            return tipo, "Formato YYYY-MM-DD"
        if tipo == "hora":
            return tipo, "Formato HH:MM (24h)"
        if tipo in ("numero", "numero_entero", "numero_decimal", "entero", "decimal"):
            return tipo, "Número entero o decimal"
        if tipo in ("url", "enlace"):
            return tipo, "URL completa (https://...)"
        if tipo in ("email", "correo"):
            return tipo, "Correo electrónico válido"
        if tipo in ("texto_corto", "texto", "texto_largo", "textarea"):
            return tipo, "Texto libre"
        return tipo, "Texto libre"

    # Hoja Instrucciones
    inst = wb.create_sheet("Instrucciones")
    inst.append(["Campo (nombre interno)", "Etiqueta visible", "Tipo", "Obligatorio", "Valores aceptados / Catálogo"])
    for cell in inst[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F3F0")
    base_inst_rows = [
        ("nombre", "Nombre completo", "texto", "Sí", "Texto libre"),
        ("email", "Correo electrónico", "email", "Sí", "Email válido. Será el usuario de acceso del jurado."),
        ("telefono", "Teléfono", "texto", "No", "Ej. 3001234567"),
        ("subregiones", "Subregiones", "multi-lista", "Sí",
         f"Uno o varios separados por ; de: Subregiones  →  Ver hoja 'Catálogos'" if subreg_valores else "Texto libre, separar con ;"),
        ("perfil", "Perfil profesional", "texto_largo", "No", "Resumen de formación y experiencia"),
    ]
    for r in base_inst_rows:
        inst.append(list(r))
    for c in campos_excel:
        tipo_label, valores = _describe_field(c)
        inst.append([
            c["nombre_interno"],
            c.get("nombre_visible") or c["nombre_interno"],
            tipo_label,
            "Sí" if c.get("obligatorio") else "No",
            valores,
        ])
    inst.append(["estado", "Estado del jurado", "lista", "No",
                 "Uno de: " + " | ".join(estados_validos) if estados_validos else "Texto libre — por defecto 'Activo'"])
    # Nota especial: anexos (firma/HV/foto) no se cargan por Excel
    inst.append([])
    inst.append(["—", "Anexos del jurado (firma, hoja de vida, foto…)", "archivo", "—",
                 "Se cargan desde el perfil del jurado (Mi Perfil), no por Excel."])
    inst.column_dimensions["A"].width = 28
    inst.column_dimensions["B"].width = 32
    inst.column_dimensions["C"].width = 14
    inst.column_dimensions["D"].width = 12
    inst.column_dimensions["E"].width = 70

    # Hoja Catálogos
    cat_sheet = wb.create_sheet("Catálogos")
    cat_sheet.append(["Catálogo", "Valor", "Descripción / Padre"])
    for cell in cat_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14776A")
    cat_ids_usados = set()
    for c in campos_excel:
        ref = c.get("catalogo_id") or c.get("catalogo")
        if ref:
            cat_ids_usados.add(ref)
    if subreg_cat:
        cat_ids_usados.add(subreg_cat.get("id"))
    if estados_cat:
        cat_ids_usados.add(estados_cat.get("id"))
    for cat in catalogos:
        if cat.get("id") not in cat_ids_usados and cat.get("nombre") not in cat_ids_usados:
            continue
        for v in (cat.get("valores") or []):
            if v.get("activo") is False:
                continue
            cat_sheet.append([cat.get("nombre"), v.get("valor"), v.get("descripcion") or v.get("padre_valor") or ""])
    cat_sheet.column_dimensions["A"].width = 28
    cat_sheet.column_dimensions["B"].width = 36
    cat_sheet.column_dimensions["C"].width = 40
    cat_sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_jurados.xlsx"})


@router.post("/jurados-import")
async def import_jurados(convocatoria_id: str = Form(...), file: UploadFile = File(...), user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"creados": 0, "rechazados": 0, "errores": []}

    # Detectar fila técnica (nombre_interno). Soporta plantilla nueva (fila 2) y vieja (fila 1).
    def _is_tech_header(r):
        if not r:
            return False
        return (str(r[0] or "")).strip().lower() == "nombre" and (str(r[1] or "")).strip().lower() == "email"

    header_row_idx = 0
    if not _is_tech_header(rows[0]) and len(rows) > 1 and _is_tech_header(rows[1]):
        header_row_idx = 1
    headers = [str(h).strip() if h else "" for h in rows[header_row_idx]]
    data_start = header_row_idx + 1

    # Catálogo Estados de Jurado
    estados_cat = await db.catalogos.find_one(
        {"convocatoria_id": convocatoria_id, "nombre": {"$in": ["Estados de Jurado", "Estados de jurado"]}},
        {"_id": 0},
    )
    estados_validos_lower = set()
    if estados_cat:
        estados_validos_lower = {
            (v.get("valor") or "").strip().lower()
            for v in (estados_cat.get("valores") or [])
            if v.get("activo") is not False
        }

    # campos jurado configurados de la convocatoria
    campos_jurado = await db.campos.find({"convocatoria_id": convocatoria_id, "aplica_a": "jurado"}, {"_id": 0}).to_list(200)
    campos_by_nombre = {c["nombre_interno"]: c for c in campos_jurado}
    BASE_KEYS = {"nombre", "email", "telefono", "subregiones", "perfil", "especialidad", "linea_experiencia", "territorio", "estado"}
    created, errors = 0, []
    for idx, row in enumerate(rows[data_start:], start=data_start + 1):
        try:
            if not any(c for c in row if c not in (None, "", " ")):
                continue
            data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            if not data.get("email") or not data.get("nombre"):
                errors.append({"fila": idx, "error": "Falta nombre o email"})
                continue

            # Validar estado contra catálogo si existe
            estado_raw = data.get("estado")
            estado = "Activo"
            if estado_raw:
                estado_str = str(estado_raw).strip()
                if estados_validos_lower:
                    if estado_str.lower() not in estados_validos_lower:
                        errors.append({"fila": idx, "error": f"Estado '{estado_str}' no es válido (ver hoja 'Catálogos')"})
                        continue
                    for v in (estados_cat.get("valores") or []):
                        if (v.get("valor") or "").strip().lower() == estado_str.lower():
                            estado = v.get("valor")
                            break
                else:
                    estado = estado_str

            # Separar base vs dinámico
            base = {k: data.get(k) for k in BASE_KEYS if data.get(k) is not None}
            datos_din = {}
            for k, v in data.items():
                if k in BASE_KEYS or not k or v is None:
                    continue
                if k in campos_by_nombre:
                    tipo = campos_by_nombre[k]["tipo"]
                    if tipo == "si_no":
                        datos_din[k] = str(v).strip().lower() in ("true", "1", "sí", "si", "yes")
                    elif tipo == "seleccion_multiple":
                        datos_din[k] = [s.strip() for s in str(v).split(";") if s.strip()]
                    else:
                        datos_din[k] = v
                else:
                    datos_din[k] = v
            # Subregiones: array si viene con ';' o ','
            subs_raw = base.get("subregiones")
            if isinstance(subs_raw, str):
                base["subregiones"] = [s.strip() for s in subs_raw.replace(",", ";").split(";") if s.strip()]
            email = str(base["email"]).strip().lower()
            existing = await db.jurados.find_one({"convocatoria_id": convocatoria_id, "email": email})
            if existing:
                errors.append({"fila": idx, "error": f"Email ya registrado: {email}"})
                continue
            jur_id = str(uuid.uuid4())
            doc = {
                "id": jur_id,
                "convocatoria_id": convocatoria_id,
                "nombre": str(base.get("nombre")).strip(),
                "email": email,
                "telefono": str(base.get("telefono") or "").strip(),
                "perfil": str(base.get("perfil") or "").strip(),
                "subregiones": base.get("subregiones") or [],
                "estado": estado,
                "disponibilidad": "Disponible",
                "datos": datos_din,
                "created_at": now_iso(),
            }
            await db.jurados.insert_one(doc)
            # Crear usuario
            existing_user = await db.users.find_one({"$or": [{"username": email}, {"email": email}]})
            if not existing_user:
                await db.users.insert_one(
                    {
                        "id": str(uuid.uuid4()),
                        "username": email,
                        "email": email,
                        "name": doc["nombre"],
                        "password_hash": hash_password("Jurado2026!"),
                        "role": "jurado",
                        "active": True,
                        "convocatoria_roles": [{"convocatoria_id": convocatoria_id, "role": "jurado"}],
                        "jurado_id": jur_id,
                        "created_at": now_iso(),
                    }
                )
            created += 1
        except Exception as e:
            errors.append({"fila": idx, "error": str(e)})
    await audit(user, "import", "jurados", convocatoria_id, detalle=f"creados={created} errores={len(errors)}")
    return {"creados": created, "rechazados": len(errors), "errores": errors[:50]}


# ==================== TERNAS ====================
class TernaIn(BaseModel):
    convocatoria_id: str
    codigo: Optional[str] = None
    nombre: str
    tipo: str = "Terna"
    integrantes: List[dict] = Field(default_factory=list)  # [{jurado_id, rol}]
    territorio: Optional[str] = None  # subregión asignada
    estado: str = "Creado"
    observaciones: Optional[str] = ""


@router.get("/ternas")
async def list_ternas(convocatoria_id: str, user: dict = Depends(get_current_user)):
    db = get_db()
    items = await db.ternas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(500)
    return items


@router.post("/ternas")
async def create_terna(payload: TernaIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    if not doc.get("codigo"):
        count = await db.ternas.count_documents({"convocatoria_id": doc["convocatoria_id"]})
        doc["codigo"] = f"T{count + 1}"
    doc["created_at"] = now_iso()
    await db.ternas.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "create", "ternas", doc["id"], valor_nuevo={"codigo": doc["codigo"]})
    return doc


@router.patch("/ternas/{tid}")
async def update_terna(tid: str, payload: dict, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    payload.pop("id", None)
    await db.ternas.update_one({"id": tid}, {"$set": payload})
    await audit(user, "update", "ternas", tid, valor_nuevo=payload)
    return await db.ternas.find_one({"id": tid}, {"_id": 0})


@router.delete("/ternas/{tid}")
async def delete_terna(tid: str, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    await db.ternas.update_one({"id": tid}, {"$set": {"estado": "Inactivo"}})
    await audit(user, "deactivate", "ternas", tid)
    return {"ok": True}


# ==================== ASIGNACIONES ====================
class AsignacionIn(BaseModel):
    convocatoria_id: str
    propuesta_id: str
    jurado_id: Optional[str] = None
    terna_id: Optional[str] = None
    tipo_evaluacion: str = "individual"  # individual | colectiva
    etapa: str = "Evaluación Individual"
    fecha_apertura: Optional[str] = None
    fecha_cierre: Optional[str] = None
    observacion: Optional[str] = ""


@router.get("/asignaciones")
async def list_asignaciones(convocatoria_id: str, jurado_id: Optional[str] = None, terna_id: Optional[str] = None, propuesta_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    db = get_db()
    q = {"convocatoria_id": convocatoria_id}
    if jurado_id:
        q["jurado_id"] = jurado_id
    if terna_id:
        q["terna_id"] = terna_id
    if propuesta_id:
        q["propuesta_id"] = propuesta_id
    items = await db.asignaciones.find(q, {"_id": 0}).to_list(5000)
    return items


@router.post("/asignaciones")
async def create_asignacion(payload: AsignacionIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    doc = payload.model_dump()
    if not doc.get("jurado_id") and not doc.get("terna_id"):
        raise HTTPException(status_code=400, detail="Debe especificar jurado_id o terna_id")
    # ──────────────── Prevenir duplicados (P0) ────────────────
    # No permite la MISMA propuesta + MISMO jurado/terna + MISMA etapa/tipo si la asignación
    # sigue activa (estado != Cancelada). Si fue cancelada, sí se puede reasignar.
    dup_q = {
        "convocatoria_id": doc["convocatoria_id"],
        "propuesta_id": doc["propuesta_id"],
        "tipo_evaluacion": doc["tipo_evaluacion"],
        "estado": {"$ne": "Cancelada"},
    }
    if doc.get("jurado_id"):
        dup_q["jurado_id"] = doc["jurado_id"]
    if doc.get("terna_id"):
        dup_q["terna_id"] = doc["terna_id"]
    if doc.get("etapa"):
        dup_q["etapa"] = doc["etapa"]
    existing = await db.asignaciones.find_one(dup_q)
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Ya existe una asignación activa para esta propuesta + {'jurado' if doc.get('jurado_id') else 'terna'} + etapa",
        )

    doc["id"] = str(uuid.uuid4())
    doc["estado"] = "Creada"
    doc["created_at"] = now_iso()
    await db.asignaciones.insert_one(doc)
    doc.pop("_id", None)

    # Auto-crear evaluación individual en estado Borrador si tipo=individual
    if doc["tipo_evaluacion"] == "individual" and doc.get("jurado_id"):
        eval_id = str(uuid.uuid4())
        await db.evaluaciones_individuales.insert_one(
            {
                "id": eval_id,
                "convocatoria_id": doc["convocatoria_id"],
                "propuesta_id": doc["propuesta_id"],
                "jurado_id": doc["jurado_id"],
                "asignacion_id": doc["id"],
                "estado": "Borrador",
                "puntajes": {},
                "observaciones": {},
                "observacion_final": "",
                "puntaje_total": 0,
                "puntaje_diferencial_total": 0,
                "created_at": now_iso(),
            }
        )
    await audit(user, "create", "asignaciones", doc["id"])
    return doc


@router.delete("/asignaciones/{aid}")
async def delete_asignacion(aid: str, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    asig = await db.asignaciones.find_one({"id": aid})
    if not asig:
        raise HTTPException(status_code=404, detail="Asignación no encontrada")
    await db.asignaciones.update_one({"id": aid}, {"$set": {"estado": "Cancelada"}})
    # Si tenía evaluación borrador, anularla
    await db.evaluaciones_individuales.update_many({"asignacion_id": aid, "estado": {"$in": ["Borrador", "Iniciada"]}}, {"$set": {"estado": "Anulada"}})
    await audit(user, "cancel", "asignaciones", aid)
    return {"ok": True}


class AsignacionMasivaIn(BaseModel):
    convocatoria_id: str
    terna_id: str
    subregion: str  # asigna todas las propuestas habilitadas de la subregión


@router.post("/asignaciones/masiva-subregion")
async def asignacion_masiva_subregion(payload: AsignacionMasivaIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    propuestas = await db.propuestas.find({"convocatoria_id": payload.convocatoria_id, "datos.subregion": payload.subregion, "estado": {"$nin": ["Anulada", "No habilitada"]}}).to_list(5000)
    creados = 0
    for p in propuestas:
        existing = await db.asignaciones.find_one({"convocatoria_id": payload.convocatoria_id, "propuesta_id": p["id"], "terna_id": payload.terna_id})
        if existing:
            continue
        await db.asignaciones.insert_one(
            {
                "id": str(uuid.uuid4()),
                "convocatoria_id": payload.convocatoria_id,
                "propuesta_id": p["id"],
                "terna_id": payload.terna_id,
                "tipo_evaluacion": "colectiva",
                "etapa": "Evaluación Colectiva",
                "estado": "Creada",
                "created_at": now_iso(),
            }
        )
        creados += 1
    # También crear asignaciones individuales para cada integrante de la terna
    terna = await db.ternas.find_one({"id": payload.terna_id})
    if terna:
        for p in propuestas:
            for integ in terna.get("integrantes", []):
                jid = integ.get("jurado_id")
                if not jid:
                    continue
                if await db.asignaciones.find_one({"convocatoria_id": payload.convocatoria_id, "propuesta_id": p["id"], "jurado_id": jid, "tipo_evaluacion": "individual"}):
                    continue
                aid = str(uuid.uuid4())
                await db.asignaciones.insert_one(
                    {
                        "id": aid,
                        "convocatoria_id": payload.convocatoria_id,
                        "propuesta_id": p["id"],
                        "jurado_id": jid,
                        "terna_id": payload.terna_id,
                        "tipo_evaluacion": "individual",
                        "etapa": "Evaluación Individual",
                        "estado": "Creada",
                        "created_at": now_iso(),
                    }
                )
                await db.evaluaciones_individuales.insert_one(
                    {
                        "id": str(uuid.uuid4()),
                        "convocatoria_id": payload.convocatoria_id,
                        "propuesta_id": p["id"],
                        "jurado_id": jid,
                        "asignacion_id": aid,
                        "estado": "Borrador",
                        "puntajes": {},
                        "observaciones": {},
                        "observacion_final": "",
                        "puntaje_total": 0,
                        "puntaje_diferencial_total": 0,
                        "created_at": now_iso(),
                    }
                )
    await audit(user, "bulk_assign", "asignaciones", payload.convocatoria_id, detalle=f"Terna {payload.terna_id} ↔ subregión {payload.subregion}: {creados} propuestas")
    return {"asignaciones_creadas": creados, "propuestas_alcanzadas": len(propuestas)}


# ==================== ASIGNACIONES: PLANTILLA / IMPORT MASIVO / AUTO ====================

class AsignacionBulkIn(BaseModel):
    convocatoria_id: str
    propuesta_ids: List[str]
    tipo_evaluacion: str  # individual | colectiva
    jurado_ids: Optional[List[str]] = None
    terna_id: Optional[str] = None
    etapa: Optional[str] = None  # default según tipo


@router.post("/asignaciones/bulk-create")
async def bulk_create_asignaciones(payload: AsignacionBulkIn,
                                    user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    """Crea N×M asignaciones (cartesiano propuestas × jurados o N×terna).
    Salta silenciosamente las combinaciones que ya existen (estado != Cancelada).
    """
    db = get_db()
    if not payload.propuesta_ids:
        raise HTTPException(400, "Selecciona al menos una propuesta")
    if payload.tipo_evaluacion == "individual" and not payload.jurado_ids:
        raise HTTPException(400, "Selecciona al menos un jurado")
    if payload.tipo_evaluacion == "colectiva" and not payload.terna_id:
        raise HTTPException(400, "Selecciona una terna")

    etapa = payload.etapa or ("Evaluación Colectiva" if payload.tipo_evaluacion == "colectiva" else "Evaluación Individual")
    creadas, duplicadas, erroneas = 0, 0, []

    targets: list = payload.jurado_ids if payload.tipo_evaluacion == "individual" else [payload.terna_id]
    for pid in payload.propuesta_ids:
        for tgt in targets:
            try:
                dup_q = {
                    "convocatoria_id": payload.convocatoria_id,
                    "propuesta_id": pid,
                    "tipo_evaluacion": payload.tipo_evaluacion,
                    "etapa": etapa,
                    "estado": {"$ne": "Cancelada"},
                }
                if payload.tipo_evaluacion == "individual":
                    dup_q["jurado_id"] = tgt
                else:
                    dup_q["terna_id"] = tgt
                if await db.asignaciones.find_one(dup_q):
                    duplicadas += 1
                    continue
                aid = str(uuid.uuid4())
                doc = {
                    "id": aid,
                    "convocatoria_id": payload.convocatoria_id,
                    "propuesta_id": pid,
                    "tipo_evaluacion": payload.tipo_evaluacion,
                    "etapa": etapa,
                    "estado": "Creada",
                    "created_at": now_iso(),
                }
                if payload.tipo_evaluacion == "individual":
                    doc["jurado_id"] = tgt
                else:
                    doc["terna_id"] = tgt
                await db.asignaciones.insert_one(doc)
                if payload.tipo_evaluacion == "individual":
                    await db.evaluaciones_individuales.insert_one({
                        "id": str(uuid.uuid4()),
                        "convocatoria_id": payload.convocatoria_id,
                        "propuesta_id": pid,
                        "jurado_id": tgt,
                        "asignacion_id": aid,
                        "estado": "Borrador",
                        "puntajes": {}, "observaciones": {}, "observacion_final": "",
                        "puntaje_total": 0, "puntaje_diferencial_total": 0,
                        "created_at": now_iso(),
                    })
                creadas += 1
            except Exception as e:
                erroneas.append({"propuesta_id": pid, "target": tgt, "error": str(e)})
    await audit(user, "bulk_create", "asignaciones", payload.convocatoria_id,
                detalle=f"{creadas} creadas, {duplicadas} ya existían")
    return {"creadas": creadas, "duplicadas": duplicadas, "erroneas": erroneas}


class AsignacionBulkDeleteIn(BaseModel):
    ids: List[str]


@router.post("/asignaciones/bulk-delete")
async def bulk_delete_asignaciones(payload: AsignacionBulkDeleteIn,
                                    user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    if not payload.ids:
        return {"canceladas": 0}
    res = await db.asignaciones.update_many(
        {"id": {"$in": payload.ids}},
        {"$set": {"estado": "Cancelada"}},
    )
    # Anular evaluaciones individuales en borrador
    await db.evaluaciones_individuales.update_many(
        {"asignacion_id": {"$in": payload.ids}, "estado": {"$in": ["Borrador", "Iniciada"]}},
        {"$set": {"estado": "Anulada"}},
    )
    await audit(user, "bulk_cancel", "asignaciones", "-",
                detalle=f"{res.modified_count} asignaciones canceladas")
    return {"canceladas": res.modified_count}


@router.get("/asignaciones-template")
async def asignaciones_template(convocatoria_id: str, user: dict = Depends(get_current_user)):
    """Plantilla XLSX con look & feel idéntico a Propuestas/Jurados:
    - 2 filas header (etiqueta humana + nombre técnico).
    - Fila ejemplo prellena con valores reales.
    - Hojas Instrucciones y Catálogos (Propuestas, Ternas, Jurados existentes + tipos válidos).
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    propuestas = await db.propuestas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).sort("codigo", 1).to_list(2000)
    ternas = await db.ternas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(500)
    jurados = await db.jurados.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(2000)

    ejemplo_propuesta = propuestas[0]["codigo"] if propuestas else "P-0001"
    ejemplo_terna = ternas[0].get("codigo", "T1") if ternas else "T1"
    ejemplo_jurado = jurados[0].get("email", "jurado@ejemplo.co") if jurados else "jurado@ejemplo.co"

    wb = Workbook()
    ws = wb.active
    ws.title = "Asignaciones"

    headers_label = [
        "Código propuesta *",
        "Tipo de evaluación *",
        "Código terna (solo colectiva)",
        "Email del jurado (solo individual)",
        "Etapa (opcional)",
    ]
    headers_tech = ["propuesta_codigo", "tipo_evaluacion", "terna_codigo", "jurado_email", "etapa"]

    ws.append(headers_label)
    label_font = Font(bold=True, color="FFFFFF", size=11)
    label_fill = PatternFill("solid", fgColor="14776A")
    for cell in ws[1]:
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    ws.append(headers_tech)
    intern_font = Font(italic=True, color="5E6878", size=9)
    intern_fill = PatternFill("solid", fgColor="F1F4F7")
    for cell in ws[2]:
        cell.font = intern_font
        cell.fill = intern_fill
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.append([ejemplo_propuesta, "colectiva", ejemplo_terna, "", "Evaluación Colectiva"])
    ws.append([ejemplo_propuesta, "individual", "", ejemplo_jurado, "Evaluación Individual"])

    for col_idx, h in enumerate(headers_label, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(18, min(40, len(h) + 5))
    ws.freeze_panes = "A3"

    # Hoja Instrucciones
    inst = wb.create_sheet("Instrucciones")
    inst.append(["Campo (técnico)", "Etiqueta visible", "Tipo", "Obligatorio", "Valores aceptados"])
    for cell in inst[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F3F0")
    inst_rows = [
        ("propuesta_codigo", "Código de la propuesta", "texto", "Sí", "Ver hoja 'Catálogos' → Propuestas (columna código)"),
        ("tipo_evaluacion", "Tipo de evaluación", "lista", "Sí", "individual | colectiva"),
        ("terna_codigo", "Código de la terna", "texto", "Solo si tipo=colectiva", "Ver hoja 'Catálogos' → Ternas"),
        ("jurado_email", "Email del jurado", "email", "Solo si tipo=individual", "Ver hoja 'Catálogos' → Jurados"),
        ("etapa", "Etapa del flujo", "texto", "No", "Por defecto: 'Evaluación Individual' o 'Evaluación Colectiva' según tipo"),
    ]
    for r in inst_rows:
        inst.append(list(r))
    inst.append([])
    inst.append(["Reglas", "—", "—", "—",
                 "No se permite la MISMA propuesta + MISMO jurado/terna + MISMA etapa duplicada. "
                 "Las filas duplicadas se rechazan automáticamente."])
    inst.column_dimensions["A"].width = 22
    inst.column_dimensions["B"].width = 28
    inst.column_dimensions["C"].width = 12
    inst.column_dimensions["D"].width = 20
    inst.column_dimensions["E"].width = 80

    # Hoja Catálogos: propuestas, ternas, jurados, tipos
    cat_sheet = wb.create_sheet("Catálogos")
    cat_sheet.append(["Catálogo", "Valor", "Descripción / Referencia"])
    for cell in cat_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14776A")
    # Tipos
    for v in ("individual", "colectiva"):
        cat_sheet.append(["Tipos de evaluación", v, ""])
    # Etapas
    for v in ("Evaluación Individual", "Evaluación Colectiva", "Consolidación", "Subregional"):
        cat_sheet.append(["Etapas", v, ""])
    # Propuestas
    for p in propuestas:
        cat_sheet.append(["Propuestas",
                           p.get("codigo"),
                           f'{p.get("nombre","") or ""} · {(p.get("datos") or {}).get("subregion","") or ""}'])
    # Ternas
    for t in ternas:
        cat_sheet.append(["Ternas",
                           t.get("codigo"),
                           f'{t.get("nombre","") or ""} · {t.get("subregion") or t.get("territorio","") or ""}'])
    # Jurados
    for j in jurados:
        cat_sheet.append(["Jurados",
                           j.get("email"),
                           f'{j.get("nombre","")} · {"; ".join(j.get("subregiones", []) or [])}'])
    cat_sheet.column_dimensions["A"].width = 22
    cat_sheet.column_dimensions["B"].width = 38
    cat_sheet.column_dimensions["C"].width = 60
    cat_sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=plantilla_asignaciones.xlsx"})


@router.post("/asignaciones-import")
async def import_asignaciones(convocatoria_id: str = Form(...), file: UploadFile = File(...), user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    db = get_db()
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"creados": 0, "rechazados": 0, "errores": []}

    # Detectar fila técnica: la plantilla nueva tiene etiqueta humana en fila 1 y
    # nombre_interno en fila 2; la vieja sólo trae nombre_interno en fila 1.
    def _is_tech(r):
        if not r:
            return False
        return (str(r[0] or "")).strip().lower() == "propuesta_codigo"

    header_row_idx = 0
    if not _is_tech(rows[0]) and len(rows) > 1 and _is_tech(rows[1]):
        header_row_idx = 1
    headers = [str(h).strip() if h else "" for h in rows[header_row_idx]]
    data_start = header_row_idx + 1

    propuestas = {p["codigo"]: p for p in await db.propuestas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(5000)}
    ternas = {t["codigo"]: t for t in await db.ternas.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(500)}
    jurados = {j["email"]: j for j in await db.jurados.find({"convocatoria_id": convocatoria_id}, {"_id": 0}).to_list(2000)}
    created, errors = 0, []
    for idx, row in enumerate(rows[data_start:], start=data_start + 1):
        try:
            if not any(c for c in row if c not in (None, "", " ")):
                continue
            data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            pcode = str(data.get("propuesta_codigo") or "").strip()
            if not pcode or pcode not in propuestas:
                errors.append({"fila": idx, "error": f"Propuesta no encontrada: {pcode}"})
                continue
            tipo = (data.get("tipo_evaluacion") or "individual").strip().lower()
            if tipo not in ("individual", "colectiva"):
                errors.append({"fila": idx, "error": f"Tipo inválido: {tipo} (debe ser individual o colectiva)"})
                continue
            etapa = (data.get("etapa") or ("Evaluación Colectiva" if tipo == "colectiva" else "Evaluación Individual")).strip()
            doc = {
                "id": str(uuid.uuid4()),
                "convocatoria_id": convocatoria_id,
                "propuesta_id": propuestas[pcode]["id"],
                "tipo_evaluacion": tipo,
                "etapa": etapa,
                "estado": "Creada",
                "created_at": now_iso(),
            }
            if tipo == "colectiva":
                tcode = str(data.get("terna_codigo") or "").strip()
                if not tcode or tcode not in ternas:
                    errors.append({"fila": idx, "error": f"Terna no encontrada: {tcode}"})
                    continue
                doc["terna_id"] = ternas[tcode]["id"]
            else:
                je = str(data.get("jurado_email") or "").strip().lower()
                if not je or je not in jurados:
                    errors.append({"fila": idx, "error": f"Jurado no encontrado: {je}"})
                    continue
                doc["jurado_id"] = jurados[je]["id"]
            # Check duplicado (mismo propuesta+jurado/terna+tipo+etapa, no cancelada)
            dup_q = {
                "convocatoria_id": convocatoria_id,
                "propuesta_id": doc["propuesta_id"],
                "tipo_evaluacion": tipo,
                "etapa": etapa,
                "estado": {"$ne": "Cancelada"},
            }
            if "jurado_id" in doc:
                dup_q["jurado_id"] = doc["jurado_id"]
            if "terna_id" in doc:
                dup_q["terna_id"] = doc["terna_id"]
            if await db.asignaciones.find_one(dup_q):
                errors.append({"fila": idx, "error": f"Asignación duplicada (propuesta {pcode} ya tiene esta {'terna' if tipo == 'colectiva' else 'jurado'} en {etapa})"})
                continue
            await db.asignaciones.insert_one(doc)
            if tipo == "individual":
                await db.evaluaciones_individuales.insert_one(
                    {
                        "id": str(uuid.uuid4()),
                        "convocatoria_id": convocatoria_id,
                        "propuesta_id": doc["propuesta_id"],
                        "jurado_id": doc["jurado_id"],
                        "asignacion_id": doc["id"],
                        "estado": "Borrador",
                        "puntajes": {},
                        "observaciones": {},
                        "observacion_final": "",
                        "puntaje_total": 0,
                        "puntaje_diferencial_total": 0,
                        "created_at": now_iso(),
                    }
                )
            created += 1
        except Exception as e:
            errors.append({"fila": idx, "error": str(e)})
    await audit(user, "import", "asignaciones", convocatoria_id, detalle=f"creados={created}")
    return {"creados": created, "rechazados": len(errors), "errores": errors[:100]}


class AutoAsignarIn(BaseModel):
    convocatoria_id: str
    jurados_por_propuesta: int = 3  # cuántos jurados individuales por propuesta
    asignar_ternas: bool = True
    solo_subregion: bool = True  # solo asignar jurados cuya subregión coincida con la de la propuesta (o "Todas")
    balance_carga: bool = True  # repartir equitativamente


@router.post("/asignaciones/auto")
async def auto_asignar(payload: AutoAsignarIn, user: dict = Depends(require_roles("admin_general", "admin_convocatoria"))):
    """Asignación automática con criterios:
    - 'solo_subregion': solo jurados cuya `subregiones[]` contenga la subregión de la propuesta o 'Todas las subregiones'.
    - 'balance_carga': de los candidatos elegibles toma los jurados con MENOS asignaciones actuales.
    - 'asignar_ternas': también enlaza la terna correspondiente a la subregión (si existe).
    Si una propuesta ya tiene N asignaciones individuales, NO se duplican.
    """
    db = get_db()
    propuestas = await db.propuestas.find({"convocatoria_id": payload.convocatoria_id, "estado": {"$nin": ["Anulada", "No habilitada"]}}, {"_id": 0}).to_list(5000)
    jurados = await db.jurados.find({"convocatoria_id": payload.convocatoria_id, "estado": "Activo"}, {"_id": 0}).to_list(2000)
    ternas = await db.ternas.find({"convocatoria_id": payload.convocatoria_id}, {"_id": 0}).to_list(500)
    # Mapa terna por subregión
    ternas_by_sub = {t.get("subregion"): t for t in ternas if t.get("subregion")}

    # Carga actual por jurado
    cur_load = {j["id"]: 0 for j in jurados}
    async for a in db.asignaciones.find({"convocatoria_id": payload.convocatoria_id, "tipo_evaluacion": "individual", "estado": {"$ne": "Cancelada"}}):
        if a.get("jurado_id") in cur_load:
            cur_load[a["jurado_id"]] += 1

    creados_ind, creados_col, omitidos = 0, 0, 0
    for p in propuestas:
        psub = (p.get("datos") or {}).get("subregion")
        # 1) Asignación colectiva (terna por subregión)
        if payload.asignar_ternas and psub and psub in ternas_by_sub:
            terna = ternas_by_sub[psub]
            already_col = await db.asignaciones.find_one({"convocatoria_id": payload.convocatoria_id, "propuesta_id": p["id"], "tipo_evaluacion": "colectiva"})
            if not already_col:
                await db.asignaciones.insert_one(
                    {
                        "id": str(uuid.uuid4()),
                        "convocatoria_id": payload.convocatoria_id,
                        "propuesta_id": p["id"],
                        "terna_id": terna["id"],
                        "tipo_evaluacion": "colectiva",
                        "etapa": "Evaluación Colectiva",
                        "estado": "Creada",
                        "created_at": now_iso(),
                    }
                )
                creados_col += 1

        # 2) Asignaciones individuales: cuántas faltan
        actuales = await db.asignaciones.count_documents({"convocatoria_id": payload.convocatoria_id, "propuesta_id": p["id"], "tipo_evaluacion": "individual", "estado": {"$ne": "Cancelada"}})
        faltan = max(0, payload.jurados_por_propuesta - actuales)
        if faltan == 0:
            omitidos += 1
            continue

        # Candidatos elegibles
        elegibles = []
        for j in jurados:
            jsubs = j.get("subregiones") or []
            ok = (not payload.solo_subregion) or (not psub) or (psub in jsubs) or ("Todas las subregiones" in jsubs)
            if not ok:
                continue
            # No re-asignar si ya está
            already = await db.asignaciones.find_one(
                {"convocatoria_id": payload.convocatoria_id, "propuesta_id": p["id"], "jurado_id": j["id"], "tipo_evaluacion": "individual", "estado": {"$ne": "Cancelada"}}
            )
            if already:
                continue
            elegibles.append(j)
        # Ordenar por carga creciente
        if payload.balance_carga:
            elegibles.sort(key=lambda x: cur_load.get(x["id"], 0))
        picked = elegibles[:faltan]
        for j in picked:
            aid = str(uuid.uuid4())
            await db.asignaciones.insert_one(
                {
                    "id": aid,
                    "convocatoria_id": payload.convocatoria_id,
                    "propuesta_id": p["id"],
                    "jurado_id": j["id"],
                    "tipo_evaluacion": "individual",
                    "etapa": "Evaluación Individual",
                    "estado": "Creada",
                    "created_at": now_iso(),
                }
            )
            await db.evaluaciones_individuales.insert_one(
                {
                    "id": str(uuid.uuid4()),
                    "convocatoria_id": payload.convocatoria_id,
                    "propuesta_id": p["id"],
                    "jurado_id": j["id"],
                    "asignacion_id": aid,
                    "estado": "Borrador",
                    "puntajes": {},
                    "observaciones": {},
                    "observacion_final": "",
                    "puntaje_total": 0,
                    "puntaje_diferencial_total": 0,
                    "created_at": now_iso(),
                }
            )
            cur_load[j["id"]] = cur_load.get(j["id"], 0) + 1
            creados_ind += 1
    await audit(user, "auto_assign", "asignaciones", payload.convocatoria_id, detalle=f"individuales={creados_ind} colectivas={creados_col}")
    return {
        "asignaciones_individuales": creados_ind,
        "asignaciones_colectivas": creados_col,
        "propuestas_omitidas_ya_completas": omitidos,
        "propuestas_total": len(propuestas),
        "jurados_activos": len(jurados),
    }
